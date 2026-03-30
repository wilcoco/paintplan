#!/usr/bin/env python3
"""
도장 생산계획 웹 앱
- 시스템 변수 관리
- 수요 데이터 관리 (엑셀 업로드)
- 생산계획 스케줄링
"""
import os
import io
from datetime import datetime, date
from flask import Flask, render_template, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from collections import defaultdict

app = Flask(__name__)

# Database configuration
database_url = os.environ.get('DATABASE_URL', 'sqlite:///paintplan.db')
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

db = SQLAlchemy(app)

# ============================================
# Database Models
# ============================================

class SystemConfig(db.Model):
    """시스템 변수"""
    __tablename__ = 'system_config'
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(50), unique=True, nullable=False)
    value = db.Column(db.String(200), nullable=False)
    description = db.Column(db.String(200))
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class JigGroup(db.Model):
    """지그 그룹"""
    __tablename__ = 'jig_groups'
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(10), unique=True, nullable=False)  # A, B, B2, C, ...
    name = db.Column(db.String(100), nullable=False)
    max_jigs = db.Column(db.Integer, nullable=False)
    pcs_per_jig = db.Column(db.Integer, default=1)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Item(db.Model):
    """아이템 마스터"""
    __tablename__ = 'items'
    id = db.Column(db.Integer, primary_key=True)
    car_type = db.Column(db.String(50), nullable=False)  # 차종
    item_name = db.Column(db.String(50), nullable=False)  # 아이템
    detail = db.Column(db.String(50))  # 세부
    color = db.Column(db.String(20), nullable=False)  # 컬러
    jig_group = db.Column(db.String(10))  # 지그그룹
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint('car_type', 'item_name', 'detail', 'color', name='uq_item'),
    )

class DailyDemand(db.Model):
    """일별 수요"""
    __tablename__ = 'daily_demands'
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('items.id'), nullable=False)
    demand_date = db.Column(db.Date, nullable=False)
    rotation = db.Column(db.Integer, nullable=False)  # 1-10
    quantity = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    item = db.relationship('Item', backref='demands')

    __table_args__ = (
        db.UniqueConstraint('item_id', 'demand_date', 'rotation', name='uq_demand'),
    )

class Inventory(db.Model):
    """재고"""
    __tablename__ = 'inventory'
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('items.id'), nullable=False)
    stock_date = db.Column(db.Date, nullable=False)
    quantity = db.Column(db.Integer, default=0)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    item = db.relationship('Item', backref='inventory')

    __table_args__ = (
        db.UniqueConstraint('item_id', 'stock_date', name='uq_inventory'),
    )

# ============================================
# Default Configuration
# ============================================

DEFAULT_CONFIG = {
    'HANGERS': ('140', '컨베이어 총 행어 수'),
    'JIGS_PER_HANGER': ('2', '행어당 지그 수'),
    'ROTATIONS_PER_DAY': ('10', '일일 회전 수'),
    'JIG_BUDGET_DAY': ('150', '주간 지그교체 예산'),
    'JIG_BUDGET_NIGHT': ('150', '야간 지그교체 예산'),
    'COLOR_CHANGE_LOSS': ('6', '컬러교환 손실'),
}

DEFAULT_JIG_GROUPS = [
    ('A', 'THPE STD/LDT+SP3', 100, 1),
    ('B', 'NQ5 FRT (STD+XLINE)', 100, 1),
    ('B2', 'NQ5 FRT STD 전용', 50, 1),
    ('C', 'OV1', 80, 1),
    ('D', 'JX EV FRT', 100, 1),
    ('E', 'JX CROSS', 80, 1),
    ('F', 'JX EV RR', 50, 1),
    ('G', 'AX PE', 80, 1),
    ('H', 'THPE RR', 50, 2),
    ('I', 'NQ5 RR', 70, 1),
]

def init_db():
    """Initialize database with default values"""
    db.create_all()

    # Insert default config
    for key, (value, desc) in DEFAULT_CONFIG.items():
        existing = SystemConfig.query.filter_by(key=key).first()
        if not existing:
            db.session.add(SystemConfig(key=key, value=value, description=desc))

    # Insert default jig groups
    for code, name, max_jigs, pcs in DEFAULT_JIG_GROUPS:
        existing = JigGroup.query.filter_by(code=code).first()
        if not existing:
            db.session.add(JigGroup(code=code, name=name, max_jigs=max_jigs, pcs_per_jig=pcs))

    db.session.commit()

# ============================================
# Helper Functions
# ============================================

def get_config():
    """Get all config as dict"""
    configs = SystemConfig.query.all()
    return {c.key: int(c.value) if c.value.isdigit() else c.value for c in configs}

def get_jig_inventory():
    """Get jig groups as dict"""
    groups = JigGroup.query.all()
    return {g.code: {'name': g.name, 'max_jigs': g.max_jigs, 'pcs': g.pcs_per_jig} for g in groups}

def get_grp(ct, it, det=''):
    """아이템을 지그그룹에 배정"""
    ct = ct.upper().replace(' ','').replace('\n','')
    it = it.upper().replace(' ','').replace('\n','')
    det = det.upper().replace(' ','').replace('\n','') if det else ''

    if 'TH' in ct:
        if 'STD' in it or 'LDT' in it: return 'A'
        if 'RR' in it: return 'H'
    if 'OV' in ct: return 'C'
    if 'NQ5' in ct:
        if 'FRT' in it:
            if 'STD' in it or 'STD' in det:
                return 'B2'
            else:
                return 'B'
        return 'I'
    if 'SP3' in ct: return 'A'
    if 'JX' in ct:
        if 'CROSS' in it: return 'E'
        if 'RR' in it: return 'F'
        return 'D'
    if 'AX' in ct or 'PE' in ct: return 'G'
    return None

def load_data_from_db(demand_date):
    """DB에서 수요 데이터 로드"""
    from datetime import timedelta

    if isinstance(demand_date, str):
        demand_date = datetime.strptime(demand_date, '%Y-%m-%d').date()

    d1_date = demand_date + timedelta(days=1)
    d2_date = demand_date + timedelta(days=2)

    items = []
    all_items = Item.query.all()

    for item in all_items:
        inv = Inventory.query.filter_by(item_id=item.id, stock_date=demand_date).first()
        stk = inv.quantity if inv else 0

        def get_demands(target_date):
            demands = DailyDemand.query.filter_by(item_id=item.id, demand_date=target_date).order_by(DailyDemand.rotation).all()
            result = [0] * 10
            for d in demands:
                if 1 <= d.rotation <= 10:
                    result[d.rotation - 1] = d.quantity
            return result

        d0 = get_demands(demand_date)
        d1 = get_demands(d1_date)
        d2 = get_demands(d2_date)

        items.append({
            'ct': item.car_type,
            'it': item.item_name,
            'det': item.detail or '-',
            'clr': item.color,
            'stk': stk,
            'd0': d0, 'd0t': sum(d0),
            'd1': d1, 'd1t': sum(d1),
            'd2': d2, 'd2t': sum(d2),
            'grp': item.jig_group or get_grp(item.car_type, item.item_name, item.detail or ''),
            'cur': stk,
            'prod': [0] * 10,
            'prod1': [0] * 10,
            'prod2': [0] * 10
        })

    return items

# ============================================
# Routes
# ============================================

@app.route('/')
def index():
    """메인 페이지"""
    return render_template('index.html')

@app.route('/api/config', methods=['GET'])
def api_get_config():
    """시스템 설정 조회"""
    configs = SystemConfig.query.all()
    return jsonify([{
        'key': c.key,
        'value': c.value,
        'description': c.description
    } for c in configs])

@app.route('/api/config', methods=['POST'])
def api_update_config():
    """시스템 설정 수정"""
    data = request.json
    for key, value in data.items():
        config = SystemConfig.query.filter_by(key=key).first()
        if config:
            config.value = str(value)
        else:
            db.session.add(SystemConfig(key=key, value=str(value)))
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/jig-groups', methods=['GET'])
def api_get_jig_groups():
    """지그그룹 조회"""
    groups = JigGroup.query.order_by(JigGroup.code).all()
    return jsonify([{
        'code': g.code,
        'name': g.name,
        'max_jigs': g.max_jigs,
        'max_hangers': g.max_jigs // 2,
        'pcs_per_jig': g.pcs_per_jig
    } for g in groups])

@app.route('/api/jig-groups', methods=['POST'])
def api_update_jig_group():
    """지그그룹 수정"""
    data = request.json
    group = JigGroup.query.filter_by(code=data['code']).first()
    if group:
        group.name = data.get('name', group.name)
        group.max_jigs = data.get('max_jigs', group.max_jigs)
        group.pcs_per_jig = data.get('pcs_per_jig', group.pcs_per_jig)
    else:
        db.session.add(JigGroup(
            code=data['code'],
            name=data['name'],
            max_jigs=data['max_jigs'],
            pcs_per_jig=data.get('pcs_per_jig', 1)
        ))
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/items', methods=['GET'])
def api_get_items():
    """아이템 목록 조회"""
    items = Item.query.order_by(Item.car_type, Item.item_name, Item.color).all()
    return jsonify([{
        'id': i.id,
        'car_type': i.car_type,
        'item_name': i.item_name,
        'detail': i.detail,
        'color': i.color,
        'jig_group': i.jig_group
    } for i in items])

@app.route('/api/demand/upload', methods=['POST'])
def api_upload_demand():
    """엑셀 파일로 수요 업로드"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Excel file required'}), 400

    demand_date = request.form.get('date')
    if not demand_date:
        return jsonify({'error': 'Date required'}), 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        items_created = 0
        demands_created = 0

        ct, it = None, None
        for r in range(8, 200):  # Adjust range as needed
            a, b, c, d, e, f = [ws.cell(r, i).value for i in range(1, 7)]

            # Skip subtotal rows
            if any(v and ('합계' in str(v) or '소계' in str(v)) for v in [b, d, e]):
                continue

            if a: ct = str(a).replace('\n', ' ').strip()
            if b: it = str(b).replace('\n', ' ').strip()
            if not ct or not it:
                continue

            clr = str(f).replace('\n', '').strip() if f else ''
            if not clr or clr == 'None':
                continue

            det = str(c).replace('\n', ' ').strip() if c else ''
            if not det or det == 'None':
                det = ''

            # Get or create item
            jig_grp = get_grp(ct, it, det)
            item = Item.query.filter_by(
                car_type=ct,
                item_name=it,
                detail=det,
                color=clr
            ).first()

            if not item:
                item = Item(
                    car_type=ct,
                    item_name=it,
                    detail=det,
                    color=clr,
                    jig_group=jig_grp
                )
                db.session.add(item)
                db.session.flush()
                items_created += 1

            # Read rotation demands for D0, D+1, D+2
            from datetime import timedelta
            base_date = datetime.strptime(demand_date, '%Y-%m-%d').date()

            demand_config = [
                # (day_offset, columns)
                (0, [22, 24, 26, 28, 30, 32, 34, 36, 38, 40]),   # D0
                (1, [43, 45, 47, 49, 51, 53, 55, 57, 59, 61]),   # D+1
                (2, [67, 69, 71, 73, 75, 77, 79, 81, 83, 85]),   # D+2
            ]

            for day_offset, demand_cols in demand_config:
                target_date = base_date + timedelta(days=day_offset)
                for rot, col in enumerate(demand_cols, 1):
                    qty = ws.cell(r, col).value
                    qty = int(qty) if isinstance(qty, (int, float)) else 0

                    existing = DailyDemand.query.filter_by(
                        item_id=item.id,
                        demand_date=target_date,
                        rotation=rot
                    ).first()

                    if existing:
                        existing.quantity = qty
                    else:
                        db.session.add(DailyDemand(
                            item_id=item.id,
                            demand_date=target_date,
                            rotation=rot,
                            quantity=qty
                        ))
                        demands_created += 1

            # Read inventory (columns G, H)
            g, h = ws.cell(r, 7).value, ws.cell(r, 8).value
            stk = (int(g) if isinstance(g, (int, float)) else 0) + (int(h) if isinstance(h, (int, float)) else 0)

            inv = Inventory.query.filter_by(
                item_id=item.id,
                stock_date=datetime.strptime(demand_date, '%Y-%m-%d').date()
            ).first()

            if inv:
                inv.quantity = stk
            else:
                db.session.add(Inventory(
                    item_id=item.id,
                    stock_date=datetime.strptime(demand_date, '%Y-%m-%d').date(),
                    quantity=stk
                ))

        db.session.commit()
        wb.close()

        return jsonify({
            'success': True,
            'items_created': items_created,
            'demands_created': demands_created
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/demand', methods=['GET'])
def api_get_demand():
    """수요 데이터 조회"""
    demand_date = request.args.get('date')
    if not demand_date:
        return jsonify({'error': 'Date required'}), 400

    demands = db.session.query(
        Item, DailyDemand, Inventory
    ).join(
        DailyDemand, Item.id == DailyDemand.item_id
    ).outerjoin(
        Inventory, db.and_(
            Item.id == Inventory.item_id,
            Inventory.stock_date == demand_date
        )
    ).filter(
        DailyDemand.demand_date == demand_date
    ).all()

    # Group by item
    items_dict = {}
    for item, demand, inv in demands:
        if item.id not in items_dict:
            items_dict[item.id] = {
                'car_type': item.car_type,
                'item_name': item.item_name,
                'detail': item.detail,
                'color': item.color,
                'jig_group': item.jig_group,
                'stock': inv.quantity if inv else 0,
                'demands': [0] * 10
            }
        items_dict[item.id]['demands'][demand.rotation - 1] = demand.quantity

    return jsonify(list(items_dict.values()))

@app.route('/api/schedule', methods=['POST'])
def api_schedule():
    """스케줄링 실행"""
    data = request.json
    demand_date = data.get('date')

    if not demand_date:
        return jsonify({'error': 'Date required'}), 400

    try:
        from generate_report import schedule
        items = load_data_from_db(demand_date)
        result = schedule(items)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/report', methods=['GET'])
def api_report():
    """HTML 리포트 생성"""
    demand_date = request.args.get('date')

    if not demand_date:
        return jsonify({'error': 'Date required'}), 400

    try:
        from generate_report import generate_html_report, schedule
        items = load_data_from_db(demand_date)
        result = schedule(items)
        html = generate_html_report(items, result)
        return html, 200, {'Content-Type': 'text/html; charset=utf-8'}
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================
# Main
# ============================================

@app.before_request
def ensure_db():
    """첫 요청 시 DB 초기화"""
    if not hasattr(app, '_db_initialized'):
        init_db()
        app._db_initialized = True

if __name__ == '__main__':
    with app.app_context():
        init_db()
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
