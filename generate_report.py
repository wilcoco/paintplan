#!/usr/bin/env python3
"""
D0 생산계획 웹 리포트 생성 v9.2
- Hard 제약: D0+D+1오전 재고부족 방지, 지그교체 ≤150/시프트, D0-1 지그교체=0
- Soft 제약: D+1오후/D+2 부족방지, 3일 안전재고
- 목적함수: 컬러교환 횟수 최소화
- 전략: 2-Pass 생산배분 (필수생산 → 회전내 컬러통일)
"""
import openpyxl
from collections import defaultdict
from datetime import datetime, timedelta

# ============================================
# 시스템 변수
# ============================================
JIG_INVENTORY = {
    'A': {'name': 'THPE STD/LDT+SP3', 'max_jigs': 100, 'pcs': 1},
    'B': {'name': 'NQ5 FRT (STD+XLINE)', 'max_jigs': 100, 'pcs': 1},
    'B2': {'name': 'NQ5 FRT STD 전용', 'max_jigs': 50, 'pcs': 1},
    'C': {'name': 'OV1', 'max_jigs': 80, 'pcs': 1},
    'D': {'name': 'JX EV FRT', 'max_jigs': 100, 'pcs': 1},
    'E': {'name': 'JX CROSS', 'max_jigs': 80, 'pcs': 1},
    'F': {'name': 'JX EV RR', 'max_jigs': 50, 'pcs': 1},
    'G': {'name': 'AX PE', 'max_jigs': 80, 'pcs': 1},
    'H': {'name': 'THPE RR', 'max_jigs': 50, 'pcs': 2},
    'I': {'name': 'NQ5 RR', 'max_jigs': 70, 'pcs': 1},
}

HANGERS = 140              # 총 행어 수
JIGS_PER_HANGER = 2        # 행어당 지그 수
ROTATIONS_PER_DAY = 10     # 일일 회전 수
DAY_SHIFT_ROTATIONS = 5    # 주간 회전 수 (1-5)
NIGHT_SHIFT_ROTATIONS = 5  # 야간 회전 수 (6-10)
HANGER_BUDGET_DAY = 150    # 주간 행어 교체 예산
HANGER_BUDGET_NIGHT = 150  # 야간 행어 교체 예산
# 컬러 교환 시 빈행어 (이전 컬러 기준)
# - 특수컬러 도장 후 → 15행어 비움
# - 일반컬러 도장 후 → 1행어 비움
SAFETY_STOCK_DAYS = 3      # 안전재고 일수

def get_grp(ct, it, det=''):
    """아이템을 지그그룹에 배정"""
    ct = ct.upper().replace(' ','').replace('\n','')
    it = it.upper().replace(' ','').replace('\n','')
    det = det.upper().replace(' ','').replace('\n','') if det else ''

    if 'TH' in ct:
        if 'STD' in it or 'LDT' in it: return 'A'
        if 'RR' in it: return 'H'
    if 'OV' in ct: return 'C'
    if 'NQ5' in ct:
        if 'FRT' in it:
            if 'STD' in it or 'STD' in det:
                return 'B2'
            else:
                return 'B'
        return 'I'
    if 'SP3' in ct: return 'A'
    if 'JX' in ct:
        if 'CROSS' in it: return 'E'
        if 'RR' in it: return 'F'
        return 'D'
    if 'AX' in ct or 'PE' in ct: return 'G'
    return None

def load_data():
    wb = openpyxl.load_workbook('paint2.xlsx', data_only=True)
    ws = wb.active
    items, ct, it = [], None, None
    for r in range(8, 138):
        a,b,c,d,e,f = [ws.cell(r,i).value for i in range(1,7)]
        if any(v and ('합계' in str(v) or '소계' in str(v)) for v in [b,d,e]): continue
        if a: ct = str(a).replace('\n',' ')
        if b: it = str(b).replace('\n',' ')
        if not ct or not it: continue
        clr = str(f).replace('\n','') if f else ''
        if not clr or clr=='None': continue
        det = str(c).replace('\n',' ').strip() if c else '-'
        if not det or det=='None': det = '-'
        g,h = ws.cell(r,7).value, ws.cell(r,8).value
        stk = (int(g) if isinstance(g,(int,float)) else 0) + (int(h) if isinstance(h,(int,float)) else 0)
        d0 = [int(ws.cell(r,c).value) if isinstance(ws.cell(r,c).value,(int,float)) else 0 for c in [22,24,26,28,30,32,34,36,38,40]]
        d1 = [int(ws.cell(r,c).value) if isinstance(ws.cell(r,c).value,(int,float)) else 0 for c in [43,45,47,49,51,53,55,57,59,61]]
        d2 = [int(ws.cell(r,c).value) if isinstance(ws.cell(r,c).value,(int,float)) else 0 for c in [67,69,71,73,75,77,79,81,83,85]]
        items.append({
            'ct':ct,'it':it,'det':det,'clr':clr,'stk':stk,
            'd0':d0,'d0t':sum(d0),
            'd1':d1,'d1t':sum(d1),
            'd2':d2,'d2t':sum(d2),
            'grp':get_grp(ct,it,det),'cur':stk,
            'prod':[0]*10,
            'prod1':[0]*10,
            'prod2':[0]*10
        })
    wb.close()
    # 수요 있는 아이템만 포함 (재고 0이어도 수요 있으면 생산 필요)
    items = [x for x in items if x['d0t'] > 0 or x['d1t'] > 0 or x['d2t'] > 0 or x['stk'] > 0]
    return items


# ============================================
# 핵심 스케줄링 함수 v9.3
# ============================================
# 목적함수: 컬러교환 최소화
# v9.3: 결정론적 알고리즘 (sorted + tie-breaker)
# Hard 제약: D0 재고부족 방지, 지그교체 ≤150/시프트
# Soft 제약: D+1 부족방지 > D+2 부족방지 > 3일 안전재고
# ============================================

# 특수컬러 (15행어=30지그 비용)
SPECIAL_COLORS = {'MGG', 'T4M', 'UMA', 'ZRM', 'ISM', 'MRM'}

# 지그그룹 클러스터 (컬러 공유 기반)
# 같은 클러스터 = 같은 컬러 생산 가능성 높음
JIG_CLUSTERS = {
    'TH': ['A', 'H'],           # TH계열: 5컬러 100% 공유
    'NQ5': ['B', 'B2', 'I'],    # NQ5계열: 9컬러 공유
    'JX': ['D', 'E', 'F', 'G'], # JX/AX계열: 6컬러 공유
    'OV': ['C'],                # OV1: 독립
}
GRP_TO_CLUSTER = {g: c for c, grps in JIG_CLUSTERS.items() for g in grps}
CLUSTER_ORDER = {'TH': 0, 'NQ5': 1, 'OV': 2, 'JX': 3}  # 클러스터 배치 순서

def get_color_change_cost(from_color):
    """컬러교환 시 빈행어 수 (이전 컬러 기준)"""
    if from_color and from_color.upper() in SPECIAL_COLORS:
        return 15  # 특수컬러 도장 후 15행어 비움
    return 1  # 일반컬러 도장 후 1행어 비움

def order_to_positions(tmpl, order):
    """템플릿과 순서를 140개 위치 배열로 변환"""
    positions = []
    for g in order:
        if g in tmpl and tmpl[g] > 0:
            positions.extend([g] * tmpl[g])
    while len(positions) < HANGERS:
        positions.append(None)
    return positions[:HANGERS]


def calc_position_changes(pos1, pos2):
    """두 위치 배열 간 교체 수 계산
    pos1=None이면 0 반환 (D0 1회전: 전날 지그 그대로 사용 가정)
    """
    if not pos1:
        return 0  # 전날 지그 그대로 사용
    if not pos2:
        return 0
    return sum(1 for i in range(HANGERS) if pos1[i] != pos2[i])


def get_optimal_order_for_colors(tmpl, grp_colors, prev_last_color):
    """컬러 교환 최소화하는 지그 순서 결정
    - 이전 마지막 컬러와 같은 컬러 먼저
    - 같은 컬러끼리 묶음
    """
    # 컬러별 지그그룹 묶기
    color_groups = defaultdict(list)
    for g, clr in sorted(grp_colors.items()):
        if g in tmpl and tmpl[g] > 0:
            color_groups[clr].append(g)

    order = []
    used_colors = set()

    # 1. 이전 컬러와 같은 컬러 먼저
    if prev_last_color and prev_last_color in color_groups:
        # 행어 수 많은 순으로 정렬
        grps = sorted(color_groups[prev_last_color], key=lambda g: -tmpl.get(g, 0))
        order.extend(grps)
        used_colors.add(prev_last_color)

    # 2. 나머지 컬러들 (행어 합계 많은 순)
    remaining = [(clr, sum(tmpl.get(g, 0) for g in grps))
                 for clr, grps in sorted(color_groups.items()) if clr not in used_colors]
    remaining.sort(key=lambda x: -x[1])

    for clr, _ in remaining:
        grps = sorted(color_groups[clr], key=lambda g: -tmpl.get(g, 0))
        order.extend(grps)

    return order


def calculate_template_for_demands(items, demand_arrays):
    """수요에 맞는 최적 템플릿 계산
    demand_arrays: list of (demand_key, weight) - 예: [('d0', 1.0), ('d1', 0.5)]
    """
    # 지그그룹별 가중 수요 계산
    grp_demand = defaultdict(float)
    for x in items:
        if x['grp']:
            for demand_key, weight in demand_arrays:
                grp_demand[x['grp']] += sum(x.get(demand_key, [0]*10)) * weight

    total_demand = sum(grp_demand.values())
    if total_demand == 0:
        return {}

    # 수요 비율로 행어 배분
    tmpl = {}
    remaining = HANGERS

    # 수요 많은 순으로 배분
    sorted_grps = sorted(grp_demand.keys(), key=lambda g: -grp_demand[g])

    for g in sorted_grps:
        if remaining <= 0:
            break
        max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER
        ideal = int(HANGERS * grp_demand[g] / total_demand)
        alloc = min(max_h, max(1 if grp_demand[g] > 0 else 0, ideal), remaining)
        if alloc > 0:
            tmpl[g] = alloc
            remaining -= alloc

    # 남은 행어 배분 (pcs=2 우선)
    if remaining > 0:
        for g in sorted_grps:
            if remaining <= 0:
                break
            if grp_demand[g] > 0:
                max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER
                current = tmpl.get(g, 0)
                add = min(remaining, max_h - current)
                if add > 0:
                    # pcs=2 지그그룹 우선
                    if JIG_INVENTORY[g]['pcs'] >= 2:
                        tmpl[g] = current + add
                        remaining -= add

    # 아직 남으면 아무 그룹에나
    if remaining > 0:
        for g in sorted_grps:
            if remaining <= 0:
                break
            max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER
            current = tmpl.get(g, 0)
            add = min(remaining, max_h - current)
            if add > 0:
                tmpl[g] = current + add
                remaining -= add

    return tmpl


def try_adjust_template(base_tmpl, base_order, target_grp, delta, budget):
    """템플릿 조정 시도 (예산 내에서)
    target_grp에 delta만큼 행어 추가/제거
    Returns: (new_tmpl, new_order, changes_used) or None if not possible
    """
    new_tmpl = base_tmpl.copy()

    if delta > 0:
        # 행어 추가
        max_h = JIG_INVENTORY[target_grp]['max_jigs'] // JIGS_PER_HANGER
        current = new_tmpl.get(target_grp, 0)
        actual_add = min(delta, max_h - current)
        if actual_add <= 0:
            return None

        # 다른 그룹에서 빼기
        other_grps = [g for g in new_tmpl if g != target_grp and new_tmpl[g] > 1]
        other_grps.sort(key=lambda g: new_tmpl[g], reverse=True)

        removed = 0
        for g in other_grps:
            if removed >= actual_add:
                break
            can_remove = new_tmpl[g] - 1
            remove = min(can_remove, actual_add - removed)
            new_tmpl[g] -= remove
            removed += remove

        if removed < actual_add:
            return None

        new_tmpl[target_grp] = current + actual_add

    elif delta < 0:
        # 행어 제거
        current = new_tmpl.get(target_grp, 0)
        actual_remove = min(-delta, current - 1) if current > 1 else 0
        if actual_remove <= 0:
            return None

        new_tmpl[target_grp] = current - actual_remove

        # 다른 그룹에 추가
        other_grps = sorted(new_tmpl.keys(), key=lambda g: -new_tmpl.get(g, 0))
        added = 0
        for g in other_grps:
            if added >= actual_remove:
                break
            if g != target_grp:
                max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER
                can_add = max_h - new_tmpl.get(g, 0)
                add = min(can_add, actual_remove - added)
                new_tmpl[g] = new_tmpl.get(g, 0) + add
                added += add

    # 순서 결정 (기존 순서 유지, 새 그룹은 뒤에)
    new_order = [g for g in base_order if g in new_tmpl and new_tmpl[g] > 0]
    for g in sorted(new_tmpl.keys()):
        if g not in new_order and new_tmpl[g] > 0:
            new_order.append(g)

    # 교체 수 계산
    old_pos = order_to_positions(base_tmpl, base_order)
    new_pos = order_to_positions(new_tmpl, new_order)
    changes = calc_position_changes(old_pos, new_pos)

    if changes > budget:
        return None

    return new_tmpl, new_order, changes


def schedule_day_v2(items, day_key, demand_key, prod_key, start_stock_key,
                    prev_template=None, prev_order=None, prev_color=None):
    """
    제약조건 기반 스케줄링 v4
    - 제약1: 지그교체 150/시프트
    - 제약2: 매 회전 생산 필수
    - 목적: 컬러교환 최소화
    - 규칙: 지그그룹당 한 컬러만 (컬러변경 시 4지그 비용)
    - 우선순위: D0+D1+D2 통합 보고 빨리 부족할 아이템 먼저
    """

    # 1. 아이템별 긴급도 계산 (D0+D1+D2 30회전 중 첫 부족 시점)
    def calc_item_urgency(x):
        """아이템의 첫 부족 회전 반환 (0~29, 없으면 30)"""
        stk = x.get(start_stock_key, x['stk'])
        for day_idx, day_k in enumerate(['d0', 'd1', 'd2']):
            demands = x.get(day_k, [0]*10)
            for r in range(10):
                stk -= demands[r]
                if stk < 0:
                    return day_idx * 10 + r
        return 30  # 부족 없음

    item_urgency = {id(x): calc_item_urgency(x) for x in items}

    # 2. 지그그룹별 주 컬러 결정 (가장 긴급한 아이템의 컬러)
    def get_grp_main_color_by_urgency(grp):
        """해당 그룹에서 가장 긴급한 아이템의 컬러"""
        grp_items = [x for x in items if x['grp'] == grp]
        if not grp_items:
            return None
        # 긴급도 순 정렬
        grp_items.sort(key=lambda x: item_urgency[id(x)])
        return grp_items[0]['clr']

    def get_grp_main_color_for_day(grp):
        """하루 전체에서 지그그룹의 주 컬러 결정
        D0 커버리지 최대화: 가장 많은 D0 수요를 만족시킬 수 있는 컬러
        (재고로 이미 커버되는 것 제외)
        """
        grp_items = [x for x in items if x['grp'] == grp]
        if not grp_items:
            return None

        # 컬러별 "생산 필요량" 계산 (재고로 부족한 부분)
        color_need = defaultdict(int)

        for x in grp_items:
            clr = x['clr']
            stk = x.get(start_stock_key, x['stk'])
            d0_total = sum(x.get(day_key, [0]*10))

            # 재고로 부족한 만큼이 생산 필요량
            need = max(0, d0_total - stk)
            color_need[clr] += need

        # 생산 필요량이 가장 큰 컬러 선택 (tie-break: 알파벳순)
        if color_need and max(color_need.values()) > 0:
            return max(color_need.keys(), key=lambda c: (color_need[c], c))

        # 모두 재고로 커버되면 수요 기반
        color_demand = defaultdict(int)
        for x in grp_items:
            color_demand[x['clr']] += sum(x.get(day_key, [0]*10))
        if color_demand:
            return max(color_demand.keys(), key=lambda c: (color_demand[c], c))
        return None

    # 3. 소량 그룹 제외 (E, F, G - 수요 미미)
    EXCLUDE_GROUPS = {'E', 'F', 'G'}
    active_items = [x for x in items if x['grp'] not in EXCLUDE_GROUPS]

    # 4. 기본 템플릿 계산 (소량 그룹 제외)
    base_tmpl = calculate_template_for_demands(active_items, [(day_key, 1.0)])

    # 5. 컬러 정보 수집 (수요 기반 - 컬러교환 최소화)
    grp_colors = {}
    for g in sorted(base_tmpl.keys()):
        grp_colors[g] = get_grp_main_color_for_day(g)

    # 6. 회전별 주컬러 계산 (지그예산 활용 위해)
    rotation_main_colors = []
    for r in range(10):
        color_demand = defaultdict(int)
        for x in active_items:
            color_demand[x['clr']] += x.get(day_key, [0]*10)[r]
        if color_demand:
            # tie-break: 알파벳순 (determinism)
            rotation_main_colors.append(max(color_demand.keys(), key=lambda c: (color_demand[c], c)))
        else:
            rotation_main_colors.append(None)

    # 5. 컬러 블록 순서 결정
    def get_color_block_order(tmpl, grp_colors, prev_last_color=None):
        """컬러 블록 단위로 순서 결정"""
        color_hangers = defaultdict(int)
        for g, clr in sorted(grp_colors.items()):
            if g in tmpl and tmpl[g] > 0 and clr:
                color_hangers[clr] += tmpl[g]

        sorted_colors = sorted(color_hangers.keys(), key=lambda c: -color_hangers[c])

        if prev_last_color and prev_last_color in sorted_colors:
            sorted_colors.remove(prev_last_color)
            sorted_colors.insert(0, prev_last_color)

        order = []
        for clr in sorted_colors:
            clr_grps = [g for g in sorted(grp_colors.keys()) if grp_colors[g] == clr and g in tmpl and tmpl[g] > 0]
            clr_grps.sort(key=lambda g: -tmpl[g])
            order.extend(clr_grps)

        return order

    # 시작 순서
    if prev_color:
        base_order = get_color_block_order(base_tmpl, grp_colors, prev_color)
    else:
        base_order = get_color_block_order(base_tmpl, grp_colors, None)

    # 결과 저장
    templates = []
    jig_orders = []
    jig_changes = [0] * 10

    # 시프트별 예산
    day_budget_left = HANGER_BUDGET_DAY
    night_budget_left = HANGER_BUDGET_NIGHT

    # 이전 상태
    # - D0 첫 회전: 전날 지그 그대로 사용 가정 → 지그 교체 0
    # - D+1/D+2 첫 회전: 전날 마지막 회전과 비교
    if prev_template and prev_order:
        prev_positions = order_to_positions(prev_template, prev_order)
    else:
        prev_positions = None  # D0: 지그 교체 0

    for r in range(10):
        is_day_shift = r < 5
        budget_left = day_budget_left if is_day_shift else night_budget_left

        # 회전별 최적 템플릿 계산 후 점진적 이동
        rot_main_color = rotation_main_colors[r]

        # 이 회전의 목표 템플릿 (주컬러 그룹 강화)
        target_tmpl = base_tmpl.copy()
        if rot_main_color:
            main_color_grps = sorted([g for g in target_tmpl if grp_colors.get(g) == rot_main_color])
            other_grps = sorted([g for g in target_tmpl if grp_colors.get(g) != rot_main_color])

            # 주컬러 그룹에 20행어 추가 목표
            transfer = 20
            for og in other_grps:
                if transfer <= 0:
                    break
                reduce = min(transfer, target_tmpl.get(og, 0) - 8)
                if reduce > 0:
                    target_tmpl[og] -= reduce
                    transfer -= reduce

            added = 20 - transfer
            if added > 0 and main_color_grps:
                per_grp = added // len(main_color_grps)
                for mg in main_color_grps:
                    max_h = JIG_INVENTORY[mg]['max_jigs'] // JIGS_PER_HANGER
                    add = min(per_grp, max_h - target_tmpl.get(mg, 0))
                    target_tmpl[mg] = target_tmpl.get(mg, 0) + add

        # 이전 템플릿에서 목표로 점진 이동 (예산 활용)
        if templates:
            curr_tmpl = templates[-1].copy()
            # 예산 여유에 따라 변경량 조절 (회전당 최대 30행어)
            max_hangers_per_rot = min(30, budget_left // 3)

            # 변경 필요량 계산
            changes_needed = []
            for g in sorted(set(curr_tmpl.keys()) | set(target_tmpl.keys())):
                curr_val = curr_tmpl.get(g, 0)
                target_val = target_tmpl.get(g, 0)
                if curr_val != target_val:
                    changes_needed.append((g, target_val - curr_val))

            # 변경량 제한하여 적용
            total_change = 0
            for g, diff in sorted(changes_needed, key=lambda x: -abs(x[1])):
                if total_change >= max_hangers_per_rot:
                    break
                apply = min(abs(diff), max_hangers_per_rot - total_change, 10)  # 그룹당 최대 10
                if diff > 0:
                    curr_tmpl[g] = curr_tmpl.get(g, 0) + apply
                else:
                    curr_tmpl[g] = curr_tmpl.get(g, 0) - apply
                total_change += apply
        else:
            curr_tmpl = base_tmpl.copy()  # 첫 회전은 기본 템플릿

        # 140행어 보장 (빈행어 방지)
        total_hangers = sum(curr_tmpl.values())
        if total_hangers < HANGERS:
            deficit = HANGERS - total_hangers
            # 주컬러 그룹에 우선 배분
            main_color_grps = sorted([g for g in curr_tmpl if grp_colors.get(g) == rot_main_color])
            other_grps = sorted([g for g in curr_tmpl if grp_colors.get(g) != rot_main_color and curr_tmpl[g] > 0])
            fill_order = main_color_grps + other_grps

            for g in fill_order:
                if deficit <= 0:
                    break
                max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER
                available = max_h - curr_tmpl.get(g, 0)
                add = min(deficit, available)
                if add > 0:
                    curr_tmpl[g] = curr_tmpl.get(g, 0) + add
                    deficit -= add

        # 순서 결정: 지그교체 vs 컬러교환 트레이드오프
        if r > 0 and jig_orders:
            prev_order = jig_orders[-1]
            prev_tmpl = templates[-1]

            # 옵션1: 이전 순서 유지
            stable_order = [g for g in prev_order if g in curr_tmpl and curr_tmpl[g] > 0]
            for g in sorted(curr_tmpl.keys()):
                if g not in stable_order and curr_tmpl[g] > 0:
                    stable_order.append(g)

            # 옵션2: 컬러 최적화 순서
            color_order = get_color_block_order(curr_tmpl, grp_colors,
                          grp_colors.get(prev_order[-1]) if prev_order else None)

            # 비용 계산
            prev_positions = order_to_positions(prev_tmpl, prev_order)
            stable_positions = order_to_positions(curr_tmpl, stable_order)
            color_positions = order_to_positions(curr_tmpl, color_order)

            stable_jig_cost = calc_position_changes(prev_positions, stable_positions)
            color_jig_cost = calc_position_changes(prev_positions, color_positions)

            # 컬러교환 차이 계산 (간단 추정: 연속 컬러 변경 횟수)
            def count_color_changes_in_order(order):
                cc = 0
                prev_clr = grp_colors.get(prev_order[-1]) if prev_order else None
                for g in order:
                    curr_clr = grp_colors.get(g)
                    if prev_clr and curr_clr and prev_clr != curr_clr:
                        cc += 1
                    prev_clr = curr_clr
                return cc

            stable_color_changes = count_color_changes_in_order(stable_order)
            color_opt_changes = count_color_changes_in_order(color_order)
            color_benefit = stable_color_changes - color_opt_changes  # 절감되는 컬러교환

            # 컬러교환 최소화 우선: 예산 내면 무조건 컬러 최적화
            # 지그교체 150개 적극 활용
            if color_jig_cost <= budget_left:
                # 컬러 이득 있으면 컬러 최적화, 없으면 비용 적은 쪽
                if color_benefit > 0 or color_jig_cost <= stable_jig_cost:
                    curr_order = color_order
                else:
                    curr_order = stable_order
            else:
                curr_order = stable_order
        else:
            # 첫 회전: 컬러 블록 순서
            curr_order = get_color_block_order(curr_tmpl, grp_colors, prev_color)

        # 지그 교체 계산 (첫 회전은 0)
        curr_positions = order_to_positions(curr_tmpl, curr_order)

        if prev_positions is not None:
            changes = calc_position_changes(prev_positions, curr_positions)

            # 예산 체크
            if changes <= budget_left:
                jig_changes[r] = changes
                if is_day_shift:
                    day_budget_left -= changes
                else:
                    night_budget_left -= changes
            else:
                # 예산 초과시 이전 템플릿/순서 유지
                if templates:
                    curr_tmpl = templates[-1].copy()
                    curr_order = list(jig_orders[-1])
                    curr_positions = order_to_positions(curr_tmpl, curr_order)
                jig_changes[r] = 0
        else:
            # 첫 회전: 전환비용 0
            jig_changes[r] = 0

        templates.append(curr_tmpl)
        jig_orders.append(curr_order)
        prev_positions = curr_positions

    # 생산량 배정 - 긴급도 기반, 지그그룹당 단일 컬러
    rotation_color_detail = [{} for _ in range(10)]

    # 각 아이템의 현재 재고 추적
    item_stock = {id(x): x.get(start_stock_key, x['stk']) for x in items}

    # 회전별 지그그룹별 생산 컬러 결정
    grp_rotation_color = {}  # {(grp, rotation): color}

    for r in range(10):
        tmpl = templates[r]
        order = jig_orders[r]

        for g in order:
            if g not in tmpl or tmpl[g] == 0:
                continue

            h = tmpl[g]
            cap = h * JIGS_PER_HANGER * JIG_INVENTORY[g]['pcs']

            # 이 지그그룹의 아이템들
            grp_items = [x for x in items if x['grp'] == g]
            if not grp_items:
                continue

            # 하루 기준 주컬러 (수요 기반) - 컬러교환 최소화
            day_main_color = grp_colors.get(g)

            # 현재 재고 기준 긴급 아이템 확인
            def calc_item_urgency_now(x):
                """현재 재고 기준 첫 부족 회전"""
                stk = item_stock[id(x)]
                for day_idx, day_k in enumerate(['d0', 'd1', 'd2']):
                    demands = x.get(day_k, [0]*10)
                    start_r = r if day_idx == 0 else 0
                    for rot in range(start_r, 10):
                        if day_idx == 0 and rot < r:
                            continue
                        stk -= demands[rot]
                        if stk < 0:
                            return day_idx * 10 + rot
                return 30

            # 긴급도 순으로 정렬
            grp_items_sorted = sorted(grp_items, key=lambda x: calc_item_urgency_now(x))

            # 컬러 선택: 하루 전체 주컬러 고정 (컬러교환 최소화 최우선)
            # 목적함수가 컬러교환 최소화이므로 주컬러 변경 없음
            main_color = day_main_color

            # 주컬러 아이템이 없으면 긴급 아이템 컬러 사용
            main_color_items = [x for x in grp_items if x['clr'] == day_main_color]
            if not main_color_items:
                most_urgent = grp_items_sorted[0] if grp_items_sorted else None
                main_color = most_urgent['clr'] if most_urgent else None

            grp_rotation_color[(g, r)] = main_color

            # ============================================
            # 목적: 컬러교환 최소화
            # Hard 제약: D0 재고부족 방지
            # ============================================
            remaining_cap = cap
            clr_hangers = defaultdict(int)
            item_alloc = {id(x): 0 for x in grp_items}

            # 모든 아이템 긴급도순 정렬
            all_items_sorted = sorted(grp_items, key=lambda x: calc_item_urgency_now(x))

            # 주컬러 아이템
            main_color_items = [x for x in grp_items if x['clr'] == main_color]
            main_color_items.sort(key=lambda x: calc_item_urgency_now(x))

            # ===== Hard 제약: D0 즉시 부족 방지 (모든 컬러) =====
            for x in all_items_sorted:
                if remaining_cap <= 0:
                    break
                current_stock = item_stock[id(x)]
                # 이번 회전에 즉시 부족한 경우만
                immediate_need = max(0, x[day_key][r] - current_stock)
                if immediate_need > 0:
                    alloc = min(remaining_cap, immediate_need)
                    item_alloc[id(x)] = alloc
                    remaining_cap -= alloc

            # ===== Soft 제약: 주컬러 우선 생산 (컬러교환 최소화) =====
            # 남은 용량은 주컬러에만 배분

            # D+1 부족 방지
            if remaining_cap > 0:
                for x in main_color_items:
                    if remaining_cap <= 0:
                        break
                    d0_remaining = sum(x.get('d0', [0]*10)[r:])
                    d1_total = sum(x.get('d1', [0]*10))
                    total_need = d0_remaining + d1_total
                    projected_stock = item_stock[id(x)] + item_alloc[id(x)]
                    shortage = max(0, total_need - projected_stock)
                    if shortage > 0:
                        alloc = min(remaining_cap, shortage)
                        item_alloc[id(x)] += alloc
                        remaining_cap -= alloc

            # D+2 부족 방지
            if remaining_cap > 0:
                for x in main_color_items:
                    if remaining_cap <= 0:
                        break
                    d0_remaining = sum(x.get('d0', [0]*10)[r:])
                    d1_total = sum(x.get('d1', [0]*10))
                    d2_total = sum(x.get('d2', [0]*10))
                    total_need = d0_remaining + d1_total + d2_total
                    projected_stock = item_stock[id(x)] + item_alloc[id(x)]
                    shortage = max(0, total_need - projected_stock)
                    if shortage > 0:
                        alloc = min(remaining_cap, shortage)
                        item_alloc[id(x)] += alloc
                        remaining_cap -= alloc

            # 남은 용량 주컬러에 균등 배분
            if remaining_cap > 0 and main_color_items:
                per_item = remaining_cap // len(main_color_items)
                extra = remaining_cap % len(main_color_items)
                for i, x in enumerate(main_color_items):
                    add = per_item + (1 if i < extra else 0)
                    item_alloc[id(x)] += add
                    remaining_cap -= add

            # 생산량 적용
            for x in grp_items:
                prod = item_alloc[id(x)]
                if prod > 0:
                    x[prod_key][r] += prod
                    clr_hangers[x['clr']] += max(1, int(h * prod / cap)) if cap > 0 else 1

            # 재고 업데이트 (모든 아이템)
            for x in grp_items:
                item_stock[id(x)] = item_stock[id(x)] - x[day_key][r] + x[prod_key][r]

            rotation_color_detail[r][g] = dict(clr_hangers)

    # 부족분 보정 (용량 제한 적용)
    # 먼저 회전별 지그그룹별 사용량 계산
    rotation_grp_used = [{} for _ in range(10)]
    for r in range(10):
        for x in items:
            g = x['grp']
            if g not in rotation_grp_used[r]:
                rotation_grp_used[r][g] = 0
            rotation_grp_used[r][g] += x[prod_key][r]

    for x in items:
        stk = x.get(start_stock_key, x['stk'])
        g = x['grp']
        grp_cap = templates[0].get(g, 0) * JIGS_PER_HANGER * JIG_INVENTORY.get(g, {'pcs': 1})['pcs']

        for r in range(10):
            stk = stk - x[day_key][r] + x[prod_key][r]
            if stk < 0:
                deficit = -stk
                # 뒤에서부터 여유 용량 내에서 추가 생산
                for pr in range(r, -1, -1):
                    if deficit <= 0:
                        break
                    # 해당 회전의 지그그룹 용량 계산
                    tmpl_h = templates[pr].get(g, 0)
                    rot_cap = tmpl_h * JIGS_PER_HANGER * JIG_INVENTORY.get(g, {'pcs': 1})['pcs']
                    used = rotation_grp_used[pr].get(g, 0)
                    available = max(0, rot_cap - used)

                    if available > 0:
                        add = min(deficit, available)
                        x[prod_key][pr] += add
                        rotation_grp_used[pr][g] = used + add
                        stk += add
                        deficit -= add

    # 실제 생산 컬러 기준 순서 재정렬 (컬러교환 최소화)
    # 지그예산 여유분 활용하여 순서 최적화
    def get_actual_grp_color(grp, rotation):
        """실제 생산 기준 지그그룹 주컬러"""
        color_prod = defaultdict(int)
        for x in items:
            if x['grp'] == grp:
                color_prod[x['clr']] += x[prod_key][rotation]
        if color_prod and max(color_prod.values()) > 0:
            return max(color_prod.keys(), key=lambda c: (color_prod[c], c))
        return None

    # 각 회전의 순서를 실제 컬러 기준으로 재정렬 (지그예산 내에서)
    day_budget_used = sum(jig_changes[:5])
    night_budget_used = sum(jig_changes[5:])
    day_budget_remain = HANGER_BUDGET_DAY - day_budget_used
    night_budget_remain = HANGER_BUDGET_NIGHT - night_budget_used

    for r in range(10):
        is_day = r < 5
        budget_remain = day_budget_remain if is_day else night_budget_remain

        curr_order = jig_orders[r]
        curr_tmpl = templates[r]

        # 현재 회전의 실제 컬러 수집
        actual_colors = {}
        for g in curr_order:
            if g in curr_tmpl and curr_tmpl[g] > 0:
                actual_colors[g] = get_actual_grp_color(g, r)

        # 컬러별로 그룹 분류
        color_groups = defaultdict(list)
        for g, clr in sorted(actual_colors.items()):
            if clr:
                color_groups[clr].append(g)

        # 컬러 순서 결정 (행어 수 많은 컬러 우선)
        color_hangers = defaultdict(int)
        for g, clr in sorted(actual_colors.items()):
            if clr:
                color_hangers[clr] += curr_tmpl.get(g, 0)
        sorted_colors = sorted(color_hangers.keys(), key=lambda c: -color_hangers[c])

        # 이전 회전 마지막 컬러 연결
        if r > 0:
            prev_order = jig_orders[r-1]
            prev_tmpl = templates[r-1]
            for g in reversed(prev_order):
                if g in prev_tmpl and prev_tmpl[g] > 0:
                    prev_last = get_actual_grp_color(g, r-1)
                    if prev_last in sorted_colors:
                        sorted_colors.remove(prev_last)
                        sorted_colors.insert(0, prev_last)
                    break

        # 새 순서 생성 (컬러별 묶음)
        new_order = []
        for clr in sorted_colors:
            grps = color_groups.get(clr, [])
            grps.sort(key=lambda g: -curr_tmpl.get(g, 0))
            new_order.extend(grps)

        # 지그교체 비용 계산
        if r > 0:
            prev_pos = order_to_positions(templates[r-1], jig_orders[r-1])
            old_pos = order_to_positions(curr_tmpl, curr_order)
            new_pos = order_to_positions(curr_tmpl, new_order)

            old_cost = calc_position_changes(prev_pos, old_pos)
            new_cost = calc_position_changes(prev_pos, new_pos)
            extra_cost = new_cost - old_cost

            # 예산 내에서만 변경
            if extra_cost <= budget_remain:
                jig_orders[r] = new_order
                jig_changes[r] = new_cost
                if is_day:
                    day_budget_remain -= extra_cost
                else:
                    night_budget_remain -= extra_cost
        else:
            jig_orders[r] = new_order

    # 컬러 교환 계산 (실제 생산 기준)
    def get_grp_main_color(grp, rotation):
        """해당 회전에서 지그그룹의 주 생산 컬러 (실제 생산 기준)"""
        color_prod = defaultdict(int)
        for x in items:
            if x['grp'] == grp:
                color_prod[x['clr']] += x[prod_key][rotation]
        if color_prod and max(color_prod.values()) > 0:
            return max(color_prod.keys(), key=lambda c: (color_prod[c], c))
        return None

    def get_grp_colors_in_rotation(grp, rotation):
        """해당 회전에서 지그그룹이 생산한 모든 컬러"""
        colors = set()
        for x in items:
            if x['grp'] == grp and x[prod_key][rotation] > 0:
                colors.add(x['clr'])
        return colors

    # 컬러교환 비용 (이전 컬러 기준 - 빈행어 수)
    # 특수컬러(MGG, T4M, UMA, ZRM, ISM, MRM) 도장 후 → 15행어 비움
    # 일반컬러 도장 후 → 1행어 비움
    SPECIAL_COLORS = {'MGG', 'T4M', 'UMA', 'ZRM', 'ISM', 'MRM'}

    def get_color_change_cost(from_color):
        """이전 컬러 도장 후 빈행어 수"""
        if from_color and from_color.upper() in SPECIAL_COLORS:
            return 15  # 특수컬러 도장 후 15행어 비움
        return 1  # 일반컬러 도장 후 1행어 비움

    def count_color_changes():
        """
        컬러교환 계산:
        - 횟수: 컬러 전환 발생 횟수
        - 빈행어: 특수컬러 후 15행어, 일반컬러 후 1행어
        """
        total_changes = 0  # 컬러교환 횟수
        total_hangers = 0  # 빈행어 합계
        hangers_per_rotation = [0] * 10
        changes_per_rotation = [0] * 10
        p_color = prev_color
        prev_grp_colors = {}

        for r in range(10):
            colors_in_order = []
            curr_grp_colors = {}

            for g in jig_orders[r]:
                if g in templates[r] and templates[r][g] > 0:
                    clr = get_grp_main_color(g, r)
                    if clr:
                        colors_in_order.append(clr)
                        curr_grp_colors[g] = clr

                    # 그룹 내 다중 컬러 - 주컬러에서 다른 컬러로 전환
                    grp_colors_set = get_grp_colors_in_rotation(g, r)
                    if len(grp_colors_set) > 1:
                        extra_count = len(grp_colors_set) - 1
                        changes_per_rotation[r] += extra_count
                        hangers_per_rotation[r] += get_color_change_cost(clr) * extra_count

            if not colors_in_order:
                prev_grp_colors = curr_grp_colors
                continue

            # 회전 간 교환 - 이전 회전 마지막 컬러 도장 후 빈행어
            if p_color and colors_in_order[0] != p_color:
                changes_per_rotation[r] += 1
                hangers_per_rotation[r] += get_color_change_cost(p_color)

            # 회전 내 교환 - 이전 컬러 도장 후 빈행어
            for i in range(1, len(colors_in_order)):
                if colors_in_order[i] != colors_in_order[i-1]:
                    changes_per_rotation[r] += 1
                    hangers_per_rotation[r] += get_color_change_cost(colors_in_order[i-1])

            # 같은 지그그룹 회전간 컬러 변경 - 이전 컬러 도장 후 빈행어
            for g, clr in sorted(curr_grp_colors.items()):
                if g in prev_grp_colors and prev_grp_colors[g] != clr:
                    changes_per_rotation[r] += 1
                    hangers_per_rotation[r] += get_color_change_cost(prev_grp_colors[g])

            total_changes += changes_per_rotation[r]
            total_hangers += hangers_per_rotation[r]
            p_color = colors_in_order[-1] if colors_in_order else p_color
            prev_grp_colors = curr_grp_colors

        return total_changes, total_hangers, changes_per_rotation, hangers_per_rotation

    cc_count, cc_hangers, cc_count_per_rot, cc_hangers_per_rot = count_color_changes()

    # 회전별 주 컬러
    rotation_color = []
    for r in range(10):
        colors = defaultdict(int)
        for g in jig_orders[r]:
            if g in templates[r] and templates[r][g] > 0:
                clr = get_grp_main_color(g, r)
                if clr:
                    colors[clr] += templates[r][g]
        rotation_color.append(max(colors.keys(), key=lambda c: (colors[c], c)) if colors else None)

    last_order = jig_orders[9]
    last_color = rotation_color[9]

    return (templates, rotation_color, jig_changes, cc_count, cc_hangers,
            cc_count_per_rot, cc_hangers_per_rot,
            rotation_color_detail, jig_orders, last_order, last_color)


def schedule_d0_optimized(items, template_override=None):
    """
    D0 최적화 스케줄러 (scheduler_v10 로직 통합)
    - 컬러 블록 배칭으로 컬러교환 최소화
    - D0 필수 그룹 항상 포함

    template_override: (template_dict, order_list) 튜플 - 외부에서 템플릿 지정 시
    """
    from collections import defaultdict

    # 특수컬러 (도장 후 15행어 비움)
    SPECIAL_COLORS_LOCAL = {'MGG', 'T4M', 'UMA', 'ZRM', 'ISM', 'MRM'}

    def get_cc_cost(from_clr, to_clr):
        """컬러 전환 비용 (이전 컬러 도장 후 빈행어 수)"""
        if from_clr == to_clr:
            return 0
        # 이전 컬러(from_clr) 기준으로 빈행어 결정
        if from_clr and from_clr.upper() in SPECIAL_COLORS_LOCAL:
            return 15  # 특수컬러 도장 후 15행어 비움
        return 1  # 일반컬러 도장 후 1행어 비움

    # Phase 1: 수요 분석
    color_demand = defaultdict(lambda: {'d0': 0, 'total': 0})
    grp_color_demand = defaultdict(lambda: defaultdict(lambda: {'d0': 0, 'must': 0}))
    color_groups = defaultdict(set)

    for x in items:
        g, clr = x['grp'], x['clr']
        d0 = sum(x.get('d0', [0]*10))
        stk = x.get('stk', 0)
        color_demand[clr]['d0'] += d0
        grp_color_demand[g][clr]['d0'] += d0
        grp_color_demand[g][clr]['must'] += max(0, d0 - stk)
        color_groups[clr].add(g)

    # Phase 2: 컬러 블록 계획
    MIN_BLOCK_DEMAND = 50  # 적정 수준
    color_rotations = {}
    for clr, demand in sorted(color_demand.items()):
        if demand['d0'] >= MIN_BLOCK_DEMAND:
            cap = sum(JIG_INVENTORY[g]['max_jigs'] for g in color_groups[clr] if g in JIG_INVENTORY)
            rotations = max(1, min(10, demand['d0'] // max(1, cap)))
            color_rotations[clr] = {'rotations': rotations, 'd0': demand['d0'], 'groups': color_groups[clr]}

    sorted_colors = sorted(color_rotations.keys(), key=lambda c: -color_rotations[c]['d0'])

    # TSP 순서
    if sorted_colors:
        sequence = [sorted_colors[0]]
        remaining = list(sorted_colors[1:])  # list for determinism
        while remaining:
            last = sequence[-1]
            # 컬러교환 비용 같으면 알파벳 순으로 tie-break (determinism)
            next_c = min(remaining, key=lambda c: (get_cc_cost(last, c), c))
            sequence.append(next_c)
            remaining.remove(next_c)
    else:
        sequence = []

    # 블록 생성
    blocks = []
    curr_rot = 0
    for clr in sequence:
        rots = color_rotations[clr]['rotations']
        if curr_rot + rots > 10:
            rots = 10 - curr_rot
        if rots > 0:
            blocks.append({'color': clr, 'start': curr_rot, 'end': curr_rot + rots - 1, 'groups': color_rotations[clr]['groups']})
            curr_rot += rots
        if curr_rot >= 10:
            break
    if curr_rot < 10 and blocks:
        blocks[-1]['end'] = 9

    # 활성 그룹
    all_groups = set(x['grp'] for x in items if x['grp'])

    # D0 필수 그룹 + D+1 오전까지 필요한 그룹
    must_groups = set()
    for g in all_groups:
        # D0 필수 (D0 수요 > 재고)
        if sum(grp_color_demand[g][c]['must'] for c in grp_color_demand[g]) > 0:
            must_groups.add(g)
    # D+1 오전까지 부족 예상 그룹 추가
    for x in items:
        g = x['grp']
        if not g:
            continue
        # D0 + D+1 오전 수요가 재고보다 많으면 필수
        d0_total = sum(x['d0'])
        d1_morning = sum(x['d1'][:5])
        if d0_total + d1_morning > x['stk']:
            must_groups.add(g)

    # Phase 3: 컬러 기반 최적 템플릿 결정
    templates, orders = [], []

    # ========================================
    # Step 1: 컬러별 수요 분석
    # ========================================
    color_demand = defaultdict(int)  # 컬러별 총 D0 수요
    color_groups = defaultdict(set)  # 컬러별 생산 가능 지그그룹
    group_colors = defaultdict(set)  # 지그그룹별 생산 컬러

    for x in items:
        g, clr = x['grp'], x['clr']
        if g and clr:
            d0_total = sum(x['d0'])
            color_demand[clr] += d0_total
            color_groups[clr].add(g)
            group_colors[g].add(clr)

    # 수요 많은 컬러 순 정렬
    sorted_colors = sorted(color_demand.keys(), key=lambda c: (-color_demand[c], c))

    # ========================================
    # Step 2: 컬러 기반 템플릿 최적화
    # ========================================
    def create_color_optimized_template():
        """컬러교환 최소화를 위한 템플릿 생성"""
        template = {}
        remaining_h = HANGERS

        # 그룹별 필요 용량 (컬러 가중치 적용)
        grp_need = defaultdict(float)
        for x in items:
            g = x['grp']
            if not g:
                continue
            d0_total = sum(x['d0'])
            d1_morning = sum(x['d1'][:5])
            need = max(0, d0_total + d1_morning - x['stk'])

            # 해당 컬러의 수요 순위에 따른 가중치
            clr = x['clr']
            if clr in sorted_colors[:5]:  # 상위 5개 컬러
                weight = 1.5
            elif clr in sorted_colors[:10]:
                weight = 1.2
            else:
                weight = 1.0

            grp_need[g] += need * weight

        total_need = sum(grp_need.values())

        # 필요량 있는 그룹 (가중 필요량 순)
        must_list = sorted([g for g in all_groups if grp_need[g] > 0],
                          key=lambda g: (-grp_need[g], g))

        if not must_list:
            d0_grp_demand = defaultdict(int)
            for x in items:
                if x['grp']:
                    d0_grp_demand[x['grp']] += sum(x['d0'])
            must_list = sorted([g for g in all_groups if d0_grp_demand[g] > 0],
                              key=lambda g: (-d0_grp_demand[g], g))

        # 배분
        for g in must_list:
            if remaining_h <= 0:
                break
            max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER
            pcs = JIG_INVENTORY[g]['pcs']

            if total_need > 0:
                ideal = int(HANGERS * grp_need[g] / total_need)
            else:
                ideal = HANGERS // len(must_list) if must_list else 10

            alloc = min(max_h, max(5, ideal), remaining_h)
            if alloc > 0:
                template[g] = alloc
                remaining_h -= alloc

        # 남은 행어 배분
        if remaining_h > 0 and must_list:
            for g in must_list:
                if remaining_h <= 0:
                    break
                max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER
                current = template.get(g, 0)
                add = min(remaining_h, max_h - current)
                if add > 0:
                    template[g] = current + add
                    remaining_h -= add

        return template

    # ========================================
    # Step 3: 컬러 기반 순서 최적화
    # ========================================
    def create_color_optimized_order(template):
        """같은 컬러 그룹끼리 연속 배치"""
        if not template:
            return []

        # 각 그룹의 주요 생산 컬러 결정
        grp_main_color = {}
        for g in template:
            color_vol = defaultdict(int)
            for x in items:
                if x['grp'] == g:
                    color_vol[x['clr']] += sum(x['d0'])
            if color_vol:
                grp_main_color[g] = max(color_vol.keys(), key=lambda c: (color_vol[c], c))

        # 컬러별 그룹 묶기
        color_to_grps = defaultdict(list)
        for g, clr in grp_main_color.items():
            color_to_grps[clr].append(g)

        # 컬러 순서: 수요 많은 컬러 먼저
        color_order = sorted(color_to_grps.keys(),
                            key=lambda c: (-color_demand.get(c, 0), c))

        # 순서 생성: 컬러별로 그룹 묶어서 배치
        order = []
        for clr in color_order:
            grps = sorted(color_to_grps[clr], key=lambda g: (-template.get(g, 0), g))
            order.extend(grps)

        # 템플릿에 있지만 컬러 정보 없는 그룹 추가
        for g in sorted(template.keys()):
            if g not in order:
                order.append(g)

        return order

    # 템플릿 생성 (외부 지정 또는 자동 생성)
    if template_override:
        base_template, base_order = template_override
        # 외부 지정 시 동일 템플릿 사용
        for r in range(10):
            templates.append(base_template.copy())
            orders.append(list(base_order))
    else:
        # ========================================
        # 동적 템플릿 전략: 회전별 다른 템플릿 사용
        # - 메인 그룹 (A, B, B2, H, I): 대부분 회전
        # - 보조 그룹 (C, D, E, F, G): 필요한 회전만
        # ========================================
        MAIN_GROUPS = {'A', 'B', 'B2', 'H', 'I'}
        SUPPLEMENTARY_GROUPS = {'C', 'D', 'E', 'F', 'G'}

        # 각 그룹별 필요 생산량 계산
        grp_need = defaultdict(int)
        for x in items:
            g = x['grp']
            if not g:
                continue
            d0_total = sum(x['d0'])
            d1_12 = sum(x['d1'][:2])  # D+1 1-2회전
            need = max(0, d0_total + d1_12 - x['stk'])
            grp_need[g] += need

        # 보조 그룹별 필요 회전 수 계산
        supplementary_rotations = {}
        for g in SUPPLEMENTARY_GROUPS:
            if grp_need[g] > 0:
                max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER
                cap_per_rot = max_h * JIGS_PER_HANGER * JIG_INVENTORY[g]['pcs']
                rots_needed = (grp_need[g] + cap_per_rot - 1) // cap_per_rot
                supplementary_rotations[g] = min(rots_needed, 10)

        # 메인 템플릿 생성 (A, B, B2, H, I만)
        def create_main_template():
            template = {}
            remaining = HANGERS
            main_demand = {g: grp_need.get(g, 0) for g in MAIN_GROUPS if g in JIG_INVENTORY}
            total_demand = sum(main_demand.values()) or 1

            for g in sorted(main_demand.keys(), key=lambda x: (-main_demand[x], x)):
                if remaining <= 0:
                    break
                max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER
                ideal = max(5, int(HANGERS * main_demand[g] / total_demand))
                alloc = min(max_h, ideal, remaining)
                if alloc > 0:
                    template[g] = alloc
                    remaining -= alloc

            # 남은 행어 배분
            for g in sorted(main_demand.keys(), key=lambda x: (-main_demand[x], x)):
                if remaining <= 0:
                    break
                max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER
                add = min(remaining, max_h - template.get(g, 0))
                if add > 0:
                    template[g] = template.get(g, 0) + add
                    remaining -= add

            return template

        # 보조 그룹 포함 템플릿 생성
        def create_supplementary_template(supp_groups):
            template = {}
            remaining = HANGERS

            # 보조 그룹 먼저 배분 (필요한 만큼만)
            for g in supp_groups:
                if g not in JIG_INVENTORY:
                    continue
                max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER
                # 최소 필요 행어 (1회전 생산량)
                alloc = min(max_h, 15, remaining)  # 최대 15행어
                if alloc > 0:
                    template[g] = alloc
                    remaining -= alloc

            # 나머지는 메인 그룹으로 채움
            main_demand = {g: grp_need.get(g, 0) for g in MAIN_GROUPS if g in JIG_INVENTORY}
            total_demand = sum(main_demand.values()) or 1

            for g in sorted(main_demand.keys(), key=lambda x: (-main_demand[x], x)):
                if remaining <= 0:
                    break
                max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER
                ideal = max(5, int(remaining * main_demand[g] / total_demand))
                alloc = min(max_h, ideal, remaining)
                if alloc > 0:
                    template[g] = template.get(g, 0) + alloc
                    remaining -= alloc

            # 남은 행어 배분
            for g in sorted(main_demand.keys(), key=lambda x: (-main_demand[x], x)):
                if remaining <= 0:
                    break
                max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER
                add = min(remaining, max_h - template.get(g, 0))
                if add > 0:
                    template[g] = template.get(g, 0) + add
                    remaining -= add

            return template

        main_template = create_main_template()
        main_order = create_color_optimized_order(main_template)

        # 회전별 템플릿 할당
        # 보조 그룹이 필요한 회전 결정 (뒤에서부터 - 데드라인 여유)
        supp_rotation_assignment = {}  # rotation -> set of supplementary groups

        for g, rots_needed in sorted(supplementary_rotations.items(), key=lambda x: -x[1]):
            # 뒤쪽 회전부터 할당 (리드타임 고려)
            assigned = 0
            for r in range(9, -1, -1):
                if assigned >= rots_needed:
                    break
                if r not in supp_rotation_assignment:
                    supp_rotation_assignment[r] = set()
                supp_rotation_assignment[r].add(g)
                assigned += 1

        # 각 회전 템플릿 생성
        for r in range(10):
            if r in supp_rotation_assignment and supp_rotation_assignment[r]:
                # 보조 그룹 포함 템플릿
                supp_template = create_supplementary_template(supp_rotation_assignment[r])
                supp_order = create_color_optimized_order(supp_template)
                templates.append(supp_template)
                orders.append(supp_order)
            else:
                # 메인 템플릿
                templates.append(main_template.copy())
                orders.append(list(main_order))

    # Phase 4: 생산 배분 (컬러 통합 최적화)
    # 전략: 주컬러 우선, 비주컬러는 최소 필수만 생산하고 지연
    for x in items:
        x['prod'] = [0] * 10
    item_stock = {id(x): x.get('stk', 0) for x in items}

    rot_main_color = {}
    for b in blocks:
        for r in range(b['start'], b['end'] + 1):
            rot_main_color[r] = b['color']

    # 그룹별 용량 계산
    def get_grp_cap(g, r):
        return templates[r].get(g, 0) * JIGS_PER_HANGER * JIG_INVENTORY[g]['pcs']

    # 컬러별 블록 회전 매핑
    color_to_rotations = defaultdict(list)
    for r, clr in sorted(rot_main_color.items()):
        if clr:
            color_to_rotations[clr].append(r)

    # 아이템별 데드라인 계산 (D0 + D+1 1-2회전까지 부족 발생 회전)
    # 2회전 리드타임 적용: R회전 부족 → R-2회전까지 생산 필요
    # D+1-3부터는 D+1 생산으로 커버 가능하므로 D0 데드라인 계산에서 제외
    LEAD_TIME = 2

    def calc_deadline(x):
        """D0+D+1 1-2회전(12회전) 내 부족 발생 시점 - 리드타임 반환, 없으면 12"""
        stk = x.get('stk', 0)
        # D0 10회전
        for r in range(10):
            stk -= x['d0'][r]
            if stk < 0:
                # D0 r회전에서 부족 → r-2까지 생산 필요
                return max(0, r - LEAD_TIME)
        # D+1 1-2회전만 (D+1-3~5는 D+1 생산으로 커버)
        for r in range(2):
            stk -= x.get('d1', [0]*10)[r]
            if stk < 0:
                # D+1 r회전에서 부족 → D0 10+r-2 = 8+r까지 생산 필요
                return max(0, 10 + r - LEAD_TIME)
        return 12  # 부족 없음 (D0 10 + D+1 2회전)

    item_deadline = {id(x): calc_deadline(x) for x in items}

    # 생산 배분: 컬러 교환 최소화하면서 데드라인 준수
    # 2-Pass 접근: 1차 - 필수 생산 (긴급/데드라인), 2차 - 선생산 (회전 컬러 통일)

    # 그룹별 남은 용량 추적
    grp_remaining_cap = [{} for _ in range(10)]

    for r in range(10):
        tmpl = templates[r]
        main_clr = rot_main_color.get(r)

        # 1차 패스: 필수 생산 (긴급 + 데드라인) + 주컬러 3일치
        for g, hangers in sorted(tmpl.items()):
            cap = hangers * JIGS_PER_HANGER * JIG_INVENTORY[g]['pcs']
            grp_items = [x for x in items if x['grp'] == g]
            if not grp_items:
                grp_remaining_cap[r][g] = 0
                continue

            remaining = cap

            # 컬러별로 그룹화
            color_items = defaultdict(list)
            for x in grp_items:
                color_items[x['clr']].append(x)

            # Step 1: 주컬러 아이템 먼저 (컬러 교환 없음)
            if main_clr and main_clr in color_items:
                for x in color_items[main_clr]:
                    if remaining <= 0:
                        break
                    # 긴급 필수 (이번 회전에서 부족)
                    curr_stock = item_stock[id(x)]
                    demand = x['d0'][r]
                    if curr_stock < demand:
                        need = demand - curr_stock
                        alloc = min(remaining, need)
                        x['prod'][r] += alloc
                        remaining -= alloc

                # 주컬러 선생산 (3일치 재고 목표: D0+D+1+D+2)
                for x in color_items[main_clr]:
                    if remaining <= 0:
                        break
                    stk = item_stock[id(x)] + x['prod'][r]
                    d0_rem = sum(x['d0'][r:])
                    d1_tot = sum(x.get('d1', [0]*10))
                    d2_tot = sum(x.get('d2', [0]*10))
                    need = max(0, d0_rem + d1_tot + d2_tot - stk)
                    if need > 0:
                        alloc = min(remaining, need)
                        x['prod'][r] += alloc
                        remaining -= alloc

            # Step 2: 비주컬러 - 데드라인 임박한 것만 생산
            # 컬러 교환 비용 순으로 정렬 (일반 < 특수)
            other_colors = [c for c in color_items.keys() if c != main_clr]
            # 특수컬러 뒤로 (교환 비용 높음)
            other_colors.sort(key=lambda c: 1 if c.upper() in SPECIAL_COLORS else 0)

            colors_used = 1 if main_clr else 0  # 사용된 컬러 수

            for clr in other_colors:
                if remaining <= 0:
                    break

                clr_items = color_items[clr]
                clr_need = 0

                for x in clr_items:
                    curr_stock = item_stock[id(x)]
                    # 이번 회전까지 생산해야 데드라인 맞추는지 확인
                    deadline = item_deadline[id(x)]
                    if deadline <= r:
                        # 이미 지났으면 즉시 생산
                        demand = x['d0'][r]
                        if curr_stock < demand:
                            clr_need += demand - curr_stock
                    elif deadline <= 10 + 5:  # D+1 오전까지
                        # 데드라인까지 남은 회전에서 생산 가능한지 확인
                        # 해당 컬러의 주컬러 회전이 데드라인 전에 있으면 거기서 생산
                        future_main_rots = [fr for fr in color_to_rotations.get(clr, [])
                                           if r < fr < deadline and fr < 10]
                        if not future_main_rots:
                            # 미룰 수 없으면 지금 생산
                            demand = x['d0'][r]
                            if curr_stock < demand:
                                clr_need += demand - curr_stock

                if clr_need > 0:
                    alloc = min(remaining, clr_need)
                    # 필요량을 아이템별로 배분
                    for x in clr_items:
                        if alloc <= 0:
                            break
                        curr_stock = item_stock[id(x)]
                        demand = x['d0'][r]
                        if curr_stock < demand:
                            item_alloc = min(alloc, demand - curr_stock)
                            x['prod'][r] += item_alloc
                            alloc -= item_alloc
                            remaining -= item_alloc
                    colors_used += 1

            # Step 3: 용량 남으면 같은 컬러 계속 생산 (컬러교환 최소화)
            # 전략: 주컬러를 3일치 초과해서도 생산 → 이미 사용중인 컬러 → 인접회전 컬러 → 기타
            if remaining > 0:
                # Step 3a: 주컬러 추가 생산 (3일치 초과해도 생산 - 컬러교환 없음)
                if main_clr and main_clr in color_items:
                    for x in color_items[main_clr]:
                        if remaining <= 0:
                            break
                        # 주컬러는 수요가 있는 한 계속 생산 (최대 D0+D1+D2 합계까지)
                        stk = item_stock[id(x)] + x['prod'][r]
                        total_demand = sum(x['d0']) + sum(x.get('d1', [0]*10)) + sum(x.get('d2', [0]*10))
                        need = max(0, total_demand - stk)
                        if need > 0:
                            alloc = min(remaining, need)
                            x['prod'][r] += alloc
                            remaining -= alloc

                # Step 3b: Step 2에서 이미 생산중인 컬러의 추가 생산 (추가 컬러교환 없음)
                colors_already_used = [c for c in other_colors
                                       if any(x['prod'][r] > 0 for x in color_items[c])]
                for clr in colors_already_used:
                    if remaining <= 0:
                        break
                    for x in color_items[clr]:
                        if remaining <= 0:
                            break
                        stk = item_stock[id(x)] + x['prod'][r]
                        d0_rem = sum(x['d0'][r:])
                        d1_tot = sum(x.get('d1', [0]*10))
                        d2_tot = sum(x.get('d2', [0]*10))
                        need = max(0, d0_rem + d1_tot + d2_tot - stk)
                        if need > 0:
                            alloc = min(remaining, need)
                            x['prod'][r] += alloc
                            remaining -= alloc

                # Step 3c: 인접 회전 주컬러 우선 (컬러 연속성)
                # 이전/다음 회전 주컬러와 같은 컬러를 우선 생산
                adjacent_colors = []
                if r > 0:
                    prev_main = rot_main_color.get(r-1)
                    if prev_main and prev_main in color_items and prev_main != main_clr:
                        adjacent_colors.append(prev_main)
                if r < 9:
                    next_main = rot_main_color.get(r+1)
                    if next_main and next_main in color_items and next_main != main_clr:
                        if next_main not in adjacent_colors:
                            adjacent_colors.append(next_main)

                for clr in adjacent_colors:
                    if remaining <= 0:
                        break
                    if clr in colors_already_used:
                        continue  # 이미 처리됨
                    for x in color_items[clr]:
                        if remaining <= 0:
                            break
                        stk = item_stock[id(x)] + x['prod'][r]
                        d0_rem = sum(x['d0'][r:])
                        d1_tot = sum(x.get('d1', [0]*10))
                        d2_tot = sum(x.get('d2', [0]*10))
                        need = max(0, d0_rem + d1_tot + d2_tot - stk)
                        if need > 0:
                            alloc = min(remaining, need)
                            x['prod'][r] += alloc
                            remaining -= alloc

                # Step 3d: 용량 채우기 (D+1 오전 부족 우선, 컬러교환 최소화)
                # 우선순위: 1) D+1 오전 부족 아이템, 2) 회전 주컬러, 3) 그룹 자체 주컬러
                if remaining > 0:
                    # 먼저 D+1 오전 부족 아이템에 배분
                    all_grp_items = []
                    for clr_list in color_items.values():
                        all_grp_items.extend(clr_list)

                    for x in all_grp_items:
                        if remaining <= 0:
                            break
                        # D0 기말재고 계산 (현재까지 생산 포함)
                        d0_end = x['stk']
                        for rr in range(r+1):
                            d0_end = d0_end - x['d0'][rr] + x['prod'][rr]
                        for rr in range(r+1, 10):
                            d0_end -= x['d0'][rr]
                        # D+1 오전 부족 체크
                        stk = d0_end
                        for rr in range(5):
                            stk -= x.get('d1', [0]*10)[rr]
                        if stk < 0:
                            # D+1 오전 부족 발생 예상 - 즉시 생산
                            need = -stk
                            alloc = min(remaining, need)
                            x['prod'][r] += alloc
                            remaining -= alloc

                    # 회전 주컬러로 채우기
                    if main_clr and main_clr in color_items:
                        main_items = [x for x in color_items[main_clr]]
                        while remaining > 0 and main_items:
                            produced_any = False
                            for x in main_items:
                                if remaining <= 0:
                                    break
                                x['prod'][r] += 1
                                remaining -= 1
                                produced_any = True
                            if not produced_any:
                                break

                    # 아직 용량 남으면 그룹 자체 주컬러로 채우기
                    if remaining > 0 and color_items:
                        # 그룹 내 가장 수요 많은 컬러
                        grp_clr_demand = {}
                        for clr, clr_items in color_items.items():
                            grp_clr_demand[clr] = sum(sum(x['d0']) for x in clr_items)
                        if grp_clr_demand:
                            grp_main_clr = max(grp_clr_demand.keys(), key=lambda c: grp_clr_demand[c])
                            if grp_main_clr in color_items:
                                grp_items = [x for x in color_items[grp_main_clr]]
                                while remaining > 0 and grp_items:
                                    produced_any = False
                                    for x in grp_items:
                                        if remaining <= 0:
                                            break
                                        x['prod'][r] += 1
                                        remaining -= 1
                                        produced_any = True
                                    if not produced_any:
                                        break

            # 남은 용량 저장 (2차 패스용)
            grp_remaining_cap[r][g] = remaining

        # 2차 패스: 회전 내 다른 그룹과 컬러 통일 (컬러교환 최소화 핵심)
        # 이 회전에서 생산되고 있는 컬러들 수집
        rot_colors_used = defaultdict(int)  # color -> total production
        rot_grp_colors = defaultdict(set)   # color -> set of groups producing it
        for x in items:
            if x['prod'][r] > 0:
                rot_colors_used[x['clr']] += x['prod'][r]
                rot_grp_colors[x['clr']].add(x['grp'])

        # 남은 용량이 있는 그룹에서 회전 내 주요 컬러 추가 생산
        if rot_colors_used:
            # 우선순위: 1) 가장 많은 그룹이 생산하는 컬러, 2) 생산량 많은 컬러
            dominant_colors = sorted(rot_colors_used.keys(),
                                   key=lambda c: (-len(rot_grp_colors[c]), -rot_colors_used[c]))

            for g in sorted(tmpl.keys()):
                remaining = grp_remaining_cap[r].get(g, 0)
                if remaining <= 0:
                    continue

                grp_items = [x for x in items if x['grp'] == g]
                if not grp_items:
                    continue

                # 이 그룹의 현재 메인 컬러 확인
                grp_color_prod = defaultdict(int)
                for x in grp_items:
                    grp_color_prod[x['clr']] += x['prod'][r]
                current_main = max(grp_color_prod.keys(), key=lambda c: (grp_color_prod[c], c)) if grp_color_prod else None

                # 현재 메인 컬러가 있으면 그것만 추가 생산 (컬러교환 방지)
                if current_main and grp_color_prod[current_main] > 0:
                    for x in grp_items:
                        if x['clr'] != current_main:
                            continue
                        if remaining <= 0:
                            break
                        stk = item_stock[id(x)] + x['prod'][r]
                        d0_rem = sum(x['d0'][r:])
                        d1_tot = sum(x.get('d1', [0]*10))
                        d2_tot = sum(x.get('d2', [0]*10))
                        need = max(0, d0_rem + d1_tot + d2_tot - stk)
                        if need > 0:
                            alloc = min(remaining, need)
                            x['prod'][r] += alloc
                            remaining -= alloc

                # 그래도 남으면 회전 내 dominant 컬러 중 하나만 선택
                if remaining > 0:
                    for dom_clr in dominant_colors:
                        if remaining <= 0:
                            break
                        # 이 컬러가 다른 그룹들에서도 많이 생산되고 있으면 우선
                        dom_clr_items = [x for x in grp_items if x['clr'] == dom_clr]
                        for x in dom_clr_items:
                            if remaining <= 0:
                                break
                            stk = item_stock[id(x)] + x['prod'][r]
                            d0_rem = sum(x['d0'][r:])
                            d1_tot = sum(x.get('d1', [0]*10))
                            d2_tot = sum(x.get('d2', [0]*10))
                            need = max(0, d0_rem + d1_tot + d2_tot - stk)
                            if need > 0:
                                alloc = min(remaining, need)
                                x['prod'][r] += alloc
                                remaining -= alloc
                        # 하나의 dominant 컬러만 추가 (컬러교환 1회로 제한)
                        if any(x['prod'][r] > 0 for x in dom_clr_items):
                            break

        # 재고 업데이트
        for x in items:
            item_stock[id(x)] = item_stock[id(x)] - x['d0'][r] + x['prod'][r]

    # =============================================
    # 야간 부족분을 주간에서 미리 생산 (시프트 선행 제약)
    # - 야간(6-10회전) 부족 → 주간(1-5회전)에서 미리 생산
    # =============================================
    # 회전별 용량 사용량 계산
    rot_grp_used = [{} for _ in range(10)]
    for r in range(10):
        for x in items:
            g = x['grp']
            if g not in rot_grp_used[r]:
                rot_grp_used[r][g] = 0
            rot_grp_used[r][g] += x['prod'][r]

    # 야간 부족 체크 및 주간 선생산
    for x in items:
        g = x['grp']
        # 재고 시뮬레이션
        stk = x['stk']
        shortage_rot = -1
        shortage_amt = 0

        for r in range(10):
            stk = stk - x['d0'][r] + x['prod'][r]
            if stk < 0 and shortage_rot < 0:
                shortage_rot = r
                shortage_amt = -stk

        # 야간 부족(6-10회전)이면 주간(1-5회전)에서 미리 생산
        if shortage_rot >= 5:  # 야간 부족
            remaining_need = shortage_amt
            # 주간 회전 뒤에서부터 여유 용량 찾기
            for pr in range(4, -1, -1):  # 5,4,3,2,1회전
                if remaining_need <= 0:
                    break
                tmpl_h = templates[pr].get(g, 0)
                cap = tmpl_h * JIGS_PER_HANGER * JIG_INVENTORY[g]['pcs']
                used = rot_grp_used[pr].get(g, 0)
                available = max(0, cap - used)
                if available > 0:
                    add = min(remaining_need, available)
                    x['prod'][pr] += add
                    rot_grp_used[pr][g] = used + add
                    remaining_need -= add

    # =============================================
    # D+1 1-2회전 부족분도 D0에서 미리 생산
    # - 2회전 리드타임: D+1-3부터는 D+1 생산으로 커버 가능
    # =============================================

    # Step 1: 그룹별 D+1 1-2회전 부족량 계산
    grp_d1_morning_shortage = defaultdict(int)
    for x in items:
        g = x['grp']
        # D0 기말재고 계산
        d0_end_stk = x['stk']
        for r in range(10):
            d0_end_stk = d0_end_stk - x['d0'][r] + x['prod'][r]

        # D+1 1-2회전 부족 체크 (D+1-3~5는 D+1 생산으로 커버)
        stk = d0_end_stk
        for r in range(2):  # D+1-1, D+1-2만
            stk -= x['d1'][r]
            if stk < 0:
                grp_d1_morning_shortage[g] += -stk
                stk = 0  # 부족분 누적을 위해 리셋

    # Step 2: 템플릿에 없는 그룹에 용량 추가
    for g, shortage in grp_d1_morning_shortage.items():
        if shortage <= 0:
            continue
        # 이미 템플릿에 충분한 용량이 있는지 확인
        total_cap = 0
        total_used = 0
        for r in range(10):
            tmpl_h = templates[r].get(g, 0)
            cap = tmpl_h * JIGS_PER_HANGER * JIG_INVENTORY[g]['pcs']
            used = rot_grp_used[r].get(g, 0)
            total_cap += cap
            total_used += used

        if total_cap - total_used >= shortage:
            continue  # 충분한 여유 있음

        # 추가 필요 용량
        additional_needed = shortage - (total_cap - total_used)
        pcs = JIG_INVENTORY[g]['pcs']
        hangers_needed = (additional_needed + (JIGS_PER_HANGER * pcs - 1)) // (JIGS_PER_HANGER * pcs)
        max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER

        # 각 회전에 용량 추가 (뒤에서부터)
        for r in range(9, -1, -1):
            if hangers_needed <= 0:
                break
            current = templates[r].get(g, 0)
            can_add = min(hangers_needed, max_h - current)
            if can_add > 0:
                # 다른 그룹에서 빼기 (여유 있는 그룹)
                for other_g in sorted(templates[r].keys(), key=lambda x: -templates[r].get(x, 0)):
                    if other_g == g:
                        continue
                    if can_add <= 0:
                        break
                    other_used = rot_grp_used[r].get(other_g, 0)
                    other_cap = templates[r][other_g] * JIGS_PER_HANGER * JIG_INVENTORY[other_g]['pcs']
                    # 최소 1행어는 유지하고 사용 중인 용량 보존
                    other_min = max(1, (other_used + JIGS_PER_HANGER * JIG_INVENTORY[other_g]['pcs'] - 1) // (JIGS_PER_HANGER * JIG_INVENTORY[other_g]['pcs']))
                    can_reduce = templates[r][other_g] - other_min
                    if can_reduce > 0:
                        reduce = min(can_reduce, can_add)
                        templates[r][other_g] -= reduce
                        templates[r][g] = templates[r].get(g, 0) + reduce
                        can_add -= reduce
                        hangers_needed -= reduce

    # Step 3: 아이템별 D+1 1-2회전 부족분 추가 생산
    for x in items:
        g = x['grp']
        # D0 기말재고 계산
        d0_end_stk = x['stk']
        for r in range(10):
            d0_end_stk = d0_end_stk - x['d0'][r] + x['prod'][r]

        # D+1 1-2회전 부족 체크 (D+1-3~5는 D+1 생산 커버)
        stk = d0_end_stk
        d1_morning_shortage = 0
        for r in range(2):  # D+1-1, D+1-2만
            stk -= x['d1'][r]
            if stk < 0:
                d1_morning_shortage = max(d1_morning_shortage, -stk)

        # D+1 1-2회전 부족분을 D0에서 추가 생산
        if d1_morning_shortage > 0:
            remaining_need = d1_morning_shortage
            # D0 뒤에서부터 여유 용량 찾기
            for pr in range(9, -1, -1):
                if remaining_need <= 0:
                    break
                tmpl_h = templates[pr].get(g, 0)
                cap = tmpl_h * JIGS_PER_HANGER * JIG_INVENTORY[g]['pcs']
                used = rot_grp_used[pr].get(g, 0)
                available = max(0, cap - used)
                if available > 0:
                    add = min(remaining_need, available)
                    x['prod'][pr] += add
                    rot_grp_used[pr][g] = used + add
                    remaining_need -= add

    # 재고 재계산
    for x in items:
        item_stock[id(x)] = x['stk']
        for r in range(10):
            item_stock[id(x)] = item_stock[id(x)] - x['d0'][r] + x['prod'][r]

    # =============================================
    # D+1 1-2회전 부족 재분배 (핵심 제약조건)
    # - 같은 그룹 내에서 과잉 생산 아이템 → 부족 아이템으로 이동
    # =============================================
    def calc_d1_morning_shortage(x):
        """D0 생산만으로 D+1 1-2회전까지 부족 수량"""
        d0_end = x['stk']
        for r in range(10):
            d0_end = d0_end - x['d0'][r] + x['prod'][r]
        stk = d0_end
        max_shortage = 0
        for r in range(2):  # D+1-1, D+1-2만
            stk -= x.get('d1', [0]*10)[r]
            if stk < 0:
                max_shortage = max(max_shortage, -stk)
        return max_shortage

    def calc_d1_morning_excess(x):
        """D0 생산으로 D+1 1-2회전까지 여유 수량 (차감 가능량)"""
        d0_end = x['stk']
        for r in range(10):
            d0_end = d0_end - x['d0'][r] + x['prod'][r]
        stk = d0_end
        min_stk = d0_end
        for r in range(2):  # D+1-1, D+1-2만
            stk -= x.get('d1', [0]*10)[r]
            min_stk = min(min_stk, stk)
        return max(0, min_stk)  # 가장 낮은 시점 기준 여유

    # 그룹별로 부족/과잉 재분배
    for g in set(x['grp'] for x in items):
        grp_items = [x for x in items if x['grp'] == g]

        for iteration in range(20):  # 최대 20회 반복
            # 현재 부족/과잉 계산
            shortage_items = [(x, calc_d1_morning_shortage(x)) for x in grp_items]
            shortage_items = [(x, s) for x, s in shortage_items if s > 0]

            if not shortage_items:
                break  # 부족 없음

            excess_items = [(x, calc_d1_morning_excess(x)) for x in grp_items]
            excess_items = [(x, e) for x, e in excess_items if e > 0 and sum(x['prod']) > 0]

            if not excess_items:
                break  # 차감 가능 항목 없음

            # 가장 부족한 아이템에 가장 여유 있는 아이템에서 이동
            shortage_items.sort(key=lambda t: -t[1])
            excess_items.sort(key=lambda t: -t[1])

            shortage_x, shortage_amt = shortage_items[0]
            excess_x, excess_amt = excess_items[0]

            # 회전별로 이동 (뒤에서부터)
            moved = 0
            for r in range(9, -1, -1):
                if moved >= shortage_amt:
                    break
                if excess_x['prod'][r] > 0:
                    # 이동량 결정 (최대 shortage_amt, excess_amt, prod[r] 중 최소)
                    transfer = min(shortage_amt - moved, excess_amt, excess_x['prod'][r])
                    if transfer > 0:
                        excess_x['prod'][r] -= transfer
                        shortage_x['prod'][r] += transfer
                        moved += transfer

            if moved == 0:
                break  # 더 이상 이동 불가

    # 재고 최종 재계산
    for x in items:
        item_stock[id(x)] = x['stk']
        for r in range(10):
            item_stock[id(x)] = item_stock[id(x)] - x['d0'][r] + x['prod'][r]

    # =============================================
    # 컬러 교환 최소화를 위한 순서 최적화
    # - 템플릿 유지, 순서만 변경
    # - 같은 컬러끼리 연속 배치
    # =============================================
    def get_grp_color(grp, rot):
        """해당 회전에서 지그그룹의 주 생산 컬러"""
        clr_prod = defaultdict(int)
        for x in items:
            if x['grp'] == grp:
                clr_prod[x['clr']] += x['prod'][rot]
        if clr_prod and max(clr_prod.values()) > 0:
            return max(clr_prod.keys(), key=lambda c: (clr_prod[c], c))
        return None

    # 컬러별 그룹 순서 재정렬 (같은 컬러끼리 연속)
    def optimize_order_by_color(rot, prev_color):
        color_grps = defaultdict(list)
        for g in sorted(templates[rot].keys()):
            if templates[rot][g] > 0:
                clr = get_grp_color(g, rot)
                if clr:
                    color_grps[clr].append(g)

        if not color_grps:
            return orders[rot]

        new_order = []
        used = set()

        # 이전 컬러 먼저
        if prev_color and prev_color in color_grps:
            for g in sorted(color_grps[prev_color], key=lambda x: -templates[rot].get(x, 0)):
                new_order.append(g)
                used.add(g)

        # 나머지 (행어 많은 순)
        rest = [(c, sum(templates[rot].get(g, 0) for g in gs))
                for c, gs in sorted(color_grps.items()) if c != prev_color]
        rest.sort(key=lambda x: -x[1])

        for c, _ in rest:
            for g in sorted(color_grps[c], key=lambda x: -templates[rot].get(x, 0)):
                if g not in used:
                    new_order.append(g)
                    used.add(g)

        # 템플릿에 있지만 생산 없는 그룹도 추가
        for g in sorted(templates[rot].keys()):
            if g not in used:
                new_order.append(g)

        return new_order

    # 컬러 최적화된 순서 계산 (행어예산 150 적극 활용)
    # 전략: 컬러 교환 최소화를 위해 행어 예산을 최대한 활용

    # 시프트별 예산
    day_remain = HANGER_BUDGET_DAY
    night_remain = HANGER_BUDGET_NIGHT

    # 전체 컬러 시퀀스 최적화 (TSP 방식)
    def get_all_colors_in_rotation(rot):
        """회전에서 생산되는 모든 컬러"""
        colors = set()
        for g in templates[rot]:
            if templates[rot][g] > 0:
                clr = get_grp_color(g, rot)
                if clr:
                    colors.add(clr)
        return colors

    def calc_color_changes_for_order(order, rot, prev_color):
        """특정 순서에서 컬러 교환 횟수 계산"""
        cc = 0
        prev_c = prev_color
        for g in order:
            if g in templates[rot] and templates[rot][g] > 0:
                c = get_grp_color(g, rot)
                if c:
                    if prev_c and c != prev_c:
                        cc += get_cc_cost(prev_c, c)
                    prev_c = c
        return cc, prev_c

    def generate_color_orders(rot, prev_color):
        """여러 순서 후보 생성 (컬러 블록 기준) - 적극적 버전"""
        from itertools import permutations

        color_grps = defaultdict(list)
        for g in sorted(templates[rot].keys()):
            if templates[rot][g] > 0:
                clr = get_grp_color(g, rot)
                if clr:
                    color_grps[clr].append(g)

        if not color_grps:
            return [sorted(templates[rot].keys())]

        # 무생산 그룹
        no_prod_grps = sorted([g for g in templates[rot] if g not in sum(color_grps.values(), [])])

        colors = sorted(color_grps.keys())
        candidates = []

        # 컬러 수가 적으면 모든 순열 시도
        if len(colors) <= 4:
            for perm in permutations(colors):
                order = []
                for c in perm:
                    for g in sorted(color_grps[c], key=lambda x: -templates[rot].get(x, 0)):
                        order.append(g)
                order.extend(no_prod_grps)
                if order not in candidates:
                    candidates.append(order)
        else:
            # 컬러 많으면 휴리스틱 사용
            # 후보 1: 이전 컬러 우선
            if prev_color and prev_color in colors:
                order = []
                for g in sorted(color_grps[prev_color], key=lambda x: -templates[rot].get(x, 0)):
                    order.append(g)
                for c in colors:
                    if c != prev_color:
                        for g in sorted(color_grps[c], key=lambda x: -templates[rot].get(x, 0)):
                            order.append(g)
                order.extend(no_prod_grps)
                candidates.append(order)

            # 후보: 각 컬러를 시작점으로
            for start_color in colors:
                order = []
                for g in sorted(color_grps[start_color], key=lambda x: -templates[rot].get(x, 0)):
                    order.append(g)
                for c in colors:
                    if c != start_color:
                        for g in sorted(color_grps[c], key=lambda x: -templates[rot].get(x, 0)):
                            order.append(g)
                order.extend(no_prod_grps)
                if order not in candidates:
                    candidates.append(order)

            # 후보: 컬러 전환 비용 최소화 순서 (그리디)
            remaining = list(colors)
            if prev_color and prev_color in remaining:
                sequence = [prev_color]
                remaining.remove(prev_color)
            else:
                sequence = [remaining.pop(0)]

            while remaining:
                last = sequence[-1]
                # 컬러교환 비용 같으면 알파벳 순으로 tie-break (determinism)
                next_c = min(remaining, key=lambda c: (get_cc_cost(last, c), c))
                sequence.append(next_c)
                remaining.remove(next_c)

            order = []
            for c in sequence:
                for g in sorted(color_grps[c], key=lambda x: -templates[rot].get(x, 0)):
                    order.append(g)
            order.extend(no_prod_grps)
            if order not in candidates:
                candidates.append(order)

        return candidates if candidates else [sorted(templates[rot].keys())]

    prev_clr = None
    prev_order = None

    for r in range(10):
        is_day = r < 5
        budget = day_remain if is_day else night_remain

        # 여러 순서 후보 생성
        candidates = generate_color_orders(r, prev_clr)

        best_order = None
        best_score = float('inf')  # 낮을수록 좋음 (컬러교환 비용)
        best_cost = 0

        for cand_order in candidates:
            # 컬러 교환 비용 계산
            cc_cost, end_color = calc_color_changes_for_order(cand_order, r, prev_clr)

            # 행어 교체 비용 계산
            if prev_order:
                prev_pos = order_to_positions(templates[r-1], prev_order)
                new_pos = order_to_positions(templates[r], cand_order)
                hanger_cost = calc_position_changes(prev_pos, new_pos)
            else:
                hanger_cost = 0

            # 예산 내에서 컬러 교환 최소화 (예산 적극 활용)
            if hanger_cost <= budget:
                score = cc_cost
                # 컬러 교환이 더 적으면 무조건 선택
                # 컬러 교환이 같으면 행어 비용 많이 써도 OK (예산 활용)
                if score < best_score:
                    best_score = score
                    best_order = cand_order
                    best_cost = hanger_cost
                elif score == best_score and best_order is None:
                    best_order = cand_order
                    best_cost = hanger_cost

        # 최적 순서가 없으면 이전 순서 유지
        if best_order is None:
            if prev_order:
                best_order = list(prev_order)
                best_cost = 0
            else:
                best_order = candidates[0]
                best_cost = 0

        orders[r] = best_order
        if is_day:
            day_remain -= best_cost
        else:
            night_remain -= best_cost

        prev_order = orders[r]

        # 이번 회전 마지막 컬러
        for g in reversed(orders[r]):
            c = get_grp_color(g, r)
            if c:
                prev_clr = c
                break


    # 지그교체 계산
    jig_changes = [0] * 10
    for r in range(1, 10):
        prev_pos = order_to_positions(templates[r-1], orders[r-1])
        curr_pos = order_to_positions(templates[r], orders[r])
        jig_changes[r] = calc_position_changes(prev_pos, curr_pos)

    # 컬러교환 계산
    def get_grp_main_clr(grp, rot):
        clr_prod = defaultdict(int)
        for x in items:
            if x['grp'] == grp:
                clr_prod[x['clr']] += x['prod'][rot]
        if clr_prod and max(clr_prod.values()) > 0:
            return max(clr_prod.keys(), key=lambda c: (clr_prod[c], c))
        return None

    def get_grp_all_colors(grp, rot):
        """그룹 내 생산되는 모든 컬러 (생산량 순 정렬)"""
        clr_prod = defaultdict(int)
        for x in items:
            if x['grp'] == grp:
                if x['prod'][rot] > 0:
                    clr_prod[x['clr']] += x['prod'][rot]
        if clr_prod:
            # 생산량 많은 순, 같으면 알파벳 순
            return sorted(clr_prod.keys(), key=lambda c: (-clr_prod[c], c))
        return []

    def get_grp_colors_optimized(grp, rot, prev_color, next_color):
        """그룹 내 컬러 순서 최적화 (이전/다음 컬러 고려)"""
        clr_prod = defaultdict(int)
        for x in items:
            if x['grp'] == grp:
                if x['prod'][rot] > 0:
                    clr_prod[x['clr']] += x['prod'][rot]
        if not clr_prod:
            return []

        colors = list(clr_prod.keys())
        if len(colors) <= 1:
            return colors

        # 최적 순서 찾기: 이전 컬러와 같은 것 먼저, 다음 컬러와 같은 것 마지막
        result = []
        remaining = set(colors)

        # 이전 컬러와 같은 것 먼저
        if prev_color and prev_color in remaining:
            result.append(prev_color)
            remaining.remove(prev_color)

        # 다음 컬러와 같은 것은 마지막에 넣기 위해 분리
        last_color = None
        if next_color and next_color in remaining and next_color != prev_color:
            last_color = next_color
            remaining.remove(next_color)

        # 나머지는 생산량 순
        for c in sorted(remaining, key=lambda c: (-clr_prod[c], c)):
            result.append(c)

        # 다음 컬러와 같은 것 마지막에
        if last_color:
            result.append(last_color)

        return result

    cc_count_total = 0  # 컬러교환 횟수
    cc_hangers_total = 0  # 빈행어 합계
    cc_count_per_rot = [0] * 10
    cc_hangers_per_rot = [0] * 10
    prev_clr = None
    for r in range(10):
        colors_ord = []
        order_list = orders[r]
        for idx, g in enumerate(order_list):
            if g in templates[r] and templates[r][g] > 0:
                # 다음 그룹의 메인 컬러 확인
                next_color = None
                for next_idx in range(idx + 1, len(order_list)):
                    next_g = order_list[next_idx]
                    if next_g in templates[r] and templates[r][next_g] > 0:
                        next_color = get_grp_main_clr(next_g, r)
                        break

                # 현재 그룹의 마지막 컬러 (이전 그룹에서 온 컬러)
                current_prev = colors_ord[-1] if colors_ord else prev_clr

                # 그룹 내 컬러 순서 최적화
                grp_colors = get_grp_colors_optimized(g, r, current_prev, next_color)
                for clr in grp_colors:
                    if clr:
                        colors_ord.append(clr)
        if prev_clr and colors_ord and colors_ord[0] != prev_clr:
            cc_count_per_rot[r] += 1
            cc_hangers_per_rot[r] += get_cc_cost(prev_clr, colors_ord[0])
        for i in range(1, len(colors_ord)):
            if colors_ord[i] != colors_ord[i-1]:
                cc_count_per_rot[r] += 1
                cc_hangers_per_rot[r] += get_cc_cost(colors_ord[i-1], colors_ord[i])
        cc_count_total += cc_count_per_rot[r]
        cc_hangers_total += cc_hangers_per_rot[r]
        if colors_ord:
            prev_clr = colors_ord[-1]

    # 회전별 주컬러
    rotation_color = []
    for r in range(10):
        clrs = defaultdict(int)
        for g in orders[r]:
            if g in templates[r]:
                clr = get_grp_main_clr(g, r)
                if clr:
                    clrs[clr] += templates[r][g]
        rotation_color.append(max(clrs.keys(), key=lambda c: (clrs[c], c)) if clrs else None)

    # 컬러 디테일
    color_detail = []
    for r in range(10):
        detail = defaultdict(int)
        for x in items:
            if x['prod'][r] > 0:
                detail[x['clr']] += x['prod'][r]
        color_detail.append(dict(detail))

    return (templates, rotation_color, jig_changes, cc_count_total, cc_hangers_total,
            cc_count_per_rot, cc_hangers_per_rot,
            color_detail, orders, orders[9], rotation_color[9])


def schedule(items):
    """D0, D+1, D+2 전체 스케줄링 (최적화 버전)"""
    import copy

    # ========================================
    # 템플릿 후보 생성 및 최적 선택
    # ========================================
    def generate_template_candidates(items):
        """
        인간 방식 템플릿 후보 생성
        - 기본: NQ계열(B, B2, I)을 앞에 배치
        - 추가: TH계열(A, H), JX계열(D, E, F, G), OV(C) 조합
        """
        from collections import defaultdict
        from itertools import combinations

        candidates = []

        # 그룹별 수요 분석
        grp_demand = defaultdict(int)
        for x in items:
            g = x['grp']
            if g:
                grp_demand[g] += sum(x['d0'])

        # ========================================
        # 그룹 클러스터 정의 (인간 방식)
        # ========================================
        NQ_GROUPS = ['B', 'B2', 'I']      # NQ5 계열 (기본)
        TH_GROUPS = ['A', 'H']             # THPE 계열
        JX_GROUPS = ['D', 'E', 'F', 'G']   # JX/AX 계열
        OV_GROUPS = ['C']                  # OV1

        # ========================================
        # 템플릿 생성 헬퍼 함수
        # ========================================
        def make_template_from_groups(group_list, name):
            """주어진 그룹 목록으로 템플릿 생성 (수요 비례 배분)"""
            if not group_list:
                return None

            # 해당 그룹들의 수요
            demands = {g: grp_demand.get(g, 0) for g in group_list if g in JIG_INVENTORY}
            if not demands:
                return None

            total_demand = sum(demands.values())
            if total_demand == 0:
                # 수요 없으면 균등 배분
                total_demand = len(demands)
                demands = {g: 1 for g in demands}

            template = {}
            remaining = HANGERS

            # 수요 비례로 배분
            sorted_grps = sorted(demands.keys(), key=lambda g: (-demands[g], g))
            for g in sorted_grps:
                if remaining <= 0:
                    break
                max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER
                ideal = max(5, int(HANGERS * demands[g] / total_demand))
                alloc = min(max_h, ideal, remaining)
                if alloc > 0:
                    template[g] = alloc
                    remaining -= alloc

            # 남은 행어 배분
            for g in sorted_grps:
                if remaining <= 0:
                    break
                max_h = JIG_INVENTORY[g]['max_jigs'] // JIGS_PER_HANGER
                add = min(remaining, max_h - template.get(g, 0))
                if add > 0:
                    template[g] = template.get(g, 0) + add
                    remaining -= add

            if sum(template.values()) != HANGERS:
                return None  # 140행어 못 채우면 무효

            # 순서: NQ를 앞에, 나머지는 클러스터 순
            order = []
            # NQ 먼저
            for g in ['B', 'B2', 'I']:
                if g in template:
                    order.append(g)
            # TH
            for g in ['A', 'H']:
                if g in template and g not in order:
                    order.append(g)
            # JX/AX
            for g in ['D', 'E', 'F', 'G']:
                if g in template and g not in order:
                    order.append(g)
            # OV
            if 'C' in template and 'C' not in order:
                order.append('C')

            return (name, template, order)

        # ========================================
        # 메인 그룹만 사용 (컬러교환 최소화)
        # C, D, E, F, G는 나중에 필요 회전에만 추가
        # ========================================

        # 메인 템플릿: NQ + TH만 (A, B, B2, H, I)
        result = make_template_from_groups(NQ_GROUPS + TH_GROUPS, 'NQ_TH_MAIN')
        if result:
            candidates.append(result)

        # 아래 후보들은 컬러교환 증가하므로 제외
        # # 타입 A 변형: NQ + TH + C (OV 추가)
        # result = make_template_from_groups(NQ_GROUPS + TH_GROUPS + OV_GROUPS, 'NQ_TH_OV')
        # if result:
        #     candidates.append(result)

        # # 타입 B: NQ + JX일부 (D만)
        # result = make_template_from_groups(NQ_GROUPS + ['D'], 'NQ_D')
        # if result:
        #     candidates.append(result)

        # # 타입 C: NQ + TH + D
        # result = make_template_from_groups(NQ_GROUPS + TH_GROUPS + ['D'], 'NQ_TH_D')
        # if result:
        #     candidates.append(result)

        # 전체 그룹은 사용하지 않음 (컬러교환 과다)
        all_groups = NQ_GROUPS + TH_GROUPS  # + JX_GROUPS + OV_GROUPS 제외
        result = make_template_from_groups(all_groups, 'ALL_GROUPS')
        if result:
            candidates.append(result)

        # NQ만 (수요 집중)
        result = make_template_from_groups(NQ_GROUPS, 'NQ_ONLY')
        if result:
            candidates.append(result)

        # NQ + A (TH의 STD만)
        result = make_template_from_groups(NQ_GROUPS + ['A'], 'NQ_A')
        if result:
            candidates.append(result)

        # NQ + A + H (TH 전체)
        result = make_template_from_groups(NQ_GROUPS + ['A', 'H'], 'NQ_A_H')
        if result:
            candidates.append(result)

        # ========================================
        # 행어 배분 변형 (동일 그룹, 다른 배분)
        # ========================================

        # NQ + TH에서 A 최대화 변형
        t_a_max = {'B': 25, 'B2': 25, 'I': 35, 'A': 50, 'H': 5}  # A 최대
        if sum(t_a_max.values()) == HANGERS:
            candidates.append(('NQ_TH_Amax', t_a_max, ['B', 'B2', 'I', 'A', 'H']))

        # NQ + TH에서 I 최대화 변형
        t_i_max = {'B': 25, 'B2': 25, 'I': 35, 'A': 30, 'H': 25}  # I 최대
        if sum(t_i_max.values()) == HANGERS:
            candidates.append(('NQ_TH_Imax', t_i_max, ['B', 'B2', 'I', 'A', 'H']))

        # NQ 집중 (B2 최대)
        t_b2_max = {'B': 50, 'B2': 25, 'I': 35, 'A': 30}  # B 최대
        if sum(t_b2_max.values()) == HANGERS:
            candidates.append(('NQ_Bmax', t_b2_max, ['B', 'B2', 'I', 'A']))

        # ========================================
        # A 최대화 변형 (검증된 최적 템플릿들)
        # ========================================

        # 기존 최적 (CC=18)
        t_best = {'A': 50, 'H': 25, 'I': 35, 'B2': 25, 'B': 5}
        if sum(t_best.values()) == HANGERS:
            candidates.append(('A50_best', t_best, ['B', 'B2', 'I', 'A', 'H']))

        # A50 변형들 (CC=18 확인됨)
        t_v2 = {'A': 50, 'H': 25, 'I': 35, 'B2': 20, 'B': 10}
        if sum(t_v2.values()) == HANGERS:
            candidates.append(('A50_v2', t_v2, ['B', 'B2', 'I', 'A', 'H']))

        t_v3 = {'A': 50, 'H': 25, 'I': 30, 'B2': 25, 'B': 10}
        if sum(t_v3.values()) == HANGERS:
            candidates.append(('A50_v3', t_v3, ['B', 'B2', 'I', 'A', 'H']))

        # I 감소 변형
        t_i25 = {'A': 50, 'H': 25, 'I': 25, 'B2': 25, 'B': 15}
        if sum(t_i25.values()) == HANGERS:
            candidates.append(('A50_I25', t_i25, ['B', 'B2', 'I', 'A', 'H']))

        # H 감소 변형
        t_h15 = {'A': 50, 'H': 15, 'I': 35, 'B2': 25, 'B': 15}
        if sum(t_h15.values()) == HANGERS:
            candidates.append(('A50_H15', t_h15, ['B', 'B2', 'I', 'A', 'H']))

        return candidates


    # 템플릿 후보 생성
    template_candidates = generate_template_candidates(items)

    # 후보 0: 기본 템플릿 (schedule_d0_optimized 내부 로직)
    # 이 템플릿을 먼저 테스트
    items_base = copy.deepcopy(items)
    base_result = schedule_d0_optimized(items_base, template_override=None)
    base_cc = base_result[3]
    base_jig = base_result[2]
    base_day = sum(base_jig[:5])
    base_night = sum(base_jig[5:])

    # 각 템플릿으로 스케줄링 시도
    best_result = None
    best_cc = float('inf')
    best_template_name = None
    best_items = None

    def count_inventory_shortage(items_after_schedule):
        """D0 각 회전 시점에서 재고부족 총량 계산"""
        total_shortage = 0
        for x in items_after_schedule:
            stk = x['stk']
            for r in range(10):
                stk = stk - x['d0'][r]
                if stk < 0:
                    total_shortage += (-stk)
                    stk = 0
                stk = stk + x['prod'][r]
        return total_shortage

    def count_d1_morning_shortage(items_after_schedule):
        """D0 생산만으로 D+1 1-2회전 재고부족 총량 계산
        (2회전 리드타임: D+1-3부터는 D+1 생산으로 커버 가능)"""
        total_shortage = 0
        for x in items_after_schedule:
            d0_end = x['stk']
            for r in range(10):
                d0_end = d0_end - x['d0'][r] + x['prod'][r]
            stk = d0_end
            # D+1-1, D+1-2만 체크 (D+1-3~5는 D+1 생산으로 커버 가능)
            for r in range(2):
                stk -= x.get('d1', [0]*10)[r]
                if stk < 0:
                    total_shortage += (-stk)
                    stk = 0
        return total_shortage

    # 기본 템플릿이 예산 내이면 기준으로 설정 (부족량도 저장)
    base_shortage = count_inventory_shortage(items_base)
    base_d1_morning = count_d1_morning_shortage(items_base)
    best_d1_morning = float('inf')

    if base_day <= HANGER_BUDGET_DAY and base_night <= HANGER_BUDGET_NIGHT:
        best_result = base_result
        best_cc = base_cc
        best_d1_morning = base_d1_morning
        best_template_name = 'base_color_optimized'
        best_items = items_base

    for name, tmpl, order in template_candidates:
        # items 깊은 복사
        items_copy = copy.deepcopy(items)

        # 스케줄링 실행
        result = schedule_d0_optimized(items_copy, template_override=(tmpl, order))

        cc_count = result[3]  # cc_count_d0

        # 행어교체 예산 체크
        jig_changes = result[2]
        day_changes = sum(jig_changes[:5])
        night_changes = sum(jig_changes[5:])

        # 재고부족량 계산
        shortage_amount = count_inventory_shortage(items_copy)
        d1_morning_shortage = count_d1_morning_shortage(items_copy)

        if day_changes <= HANGER_BUDGET_DAY and night_changes <= HANGER_BUDGET_NIGHT and shortage_amount <= base_shortage:
            # 선택 기준: 1) D+1 오전 부족 최소화, 2) 컬러교환 최소화
            is_better = False
            if d1_morning_shortage < best_d1_morning:
                is_better = True  # D+1 오전 부족이 더 적음
            elif d1_morning_shortage == best_d1_morning and cc_count < best_cc:
                is_better = True  # D+1 오전 부족 동일, 컬러교환 더 적음

            if is_better:
                best_cc = cc_count
                best_d1_morning = d1_morning_shortage
                best_result = result
                best_template_name = name
                best_items = items_copy

    # 최적 결과가 없으면 기본 실행
    if best_result is None:
        (templates_d0, rotation_color_d0, jig_changes_d0, cc_count_d0, cc_hangers_d0,
         cc_count_per_rot_d0, cc_hangers_per_rot_d0,
         color_detail_d0, jig_orders_d0, d0_last_order, d0_last_color) = schedule_d0_optimized(items)
    else:
        (templates_d0, rotation_color_d0, jig_changes_d0, cc_count_d0, cc_hangers_d0,
         cc_count_per_rot_d0, cc_hangers_per_rot_d0,
         color_detail_d0, jig_orders_d0, d0_last_order, d0_last_color) = best_result

        # 원본 items에 최적 결과 적용
        for i, x in enumerate(items):
            x['prod'] = best_items[i]['prod']

    # =============================================
    # D+1 1-2회전 부족 그룹 템플릿 보완
    # - 템플릿에 없는 그룹에 최소 용량 추가
    # =============================================
    from collections import defaultdict as dd_grp
    grp_d1_morning_need = dd_grp(int)
    for x in items:
        g = x['grp']
        # D0 기말재고 계산
        d0_end = x['stk']
        for r in range(10):
            d0_end = d0_end - x['d0'][r] + x['prod'][r]
        # D+1 1-2회전 부족 계산
        stk = d0_end
        for r in range(2):  # D+1-1, D+1-2만
            stk -= x.get('d1', [0]*10)[r]
            if stk < 0:
                grp_d1_morning_need[g] += -stk
                stk = 0

    # 현재 생산량 계산 (템플릿 수정 전)
    rot_grp_production = [{} for _ in range(10)]
    for r in range(10):
        for x in items:
            g_prod = x['grp']
            if g_prod not in rot_grp_production[r]:
                rot_grp_production[r][g_prod] = 0
            rot_grp_production[r][g_prod] += x['prod'][r]

    # 템플릿에 없거나 용량 부족한 그룹에 추가
    for g, need in sorted(grp_d1_morning_need.items(), key=lambda x: -x[1]):
        if need <= 0:
            continue
        # 현재 템플릿 용량 확인
        pcs = JIG_INVENTORY.get(g, {'pcs': 1})['pcs']
        max_h = JIG_INVENTORY.get(g, {'max_jigs': 10})['max_jigs'] // JIGS_PER_HANGER
        total_cap = sum(templates_d0[r].get(g, 0) * JIGS_PER_HANGER * pcs for r in range(10))

        if total_cap >= need:
            continue  # 이미 충분

        # 필요 추가 행어 계산
        hangers_to_add = (need - total_cap + JIGS_PER_HANGER * pcs - 1) // (JIGS_PER_HANGER * pcs)
        hangers_to_add = min(hangers_to_add, max_h * 3)  # 최대 3회전 분량으로 제한

        # 특정 회전에만 분배 (뒤쪽 8,9,10회전만 - 컬러교환 최소화)
        # D+1 1-2회전 커버를 위해 D0 후반부에 집중
        target_rotations = [9, 8, 7]  # 우선순위: 10, 9, 8회전 (0-indexed)

        for r in target_rotations:
            if hangers_to_add <= 0:
                break
            current = templates_d0[r].get(g, 0)
            can_add = min(max_h - current, hangers_to_add)
            if can_add <= 0:
                continue

            # 가장 행어 많은 그룹에서 빼기 (실제 사용량 초과 금지)
            for other_g in sorted(templates_d0[r].keys(), key=lambda x: -templates_d0[r].get(x, 0)):
                if other_g == g or other_g not in templates_d0[r]:
                    continue
                if can_add <= 0:
                    break
                other_current = templates_d0[r][other_g]
                other_pcs = JIG_INVENTORY.get(other_g, {'pcs': 1})['pcs']
                # 실제 사용량 기준 최소 행어 계산
                other_used = rot_grp_production[r].get(other_g, 0)
                other_min_h = (other_used + JIGS_PER_HANGER * other_pcs - 1) // (JIGS_PER_HANGER * other_pcs)
                other_min_h = max(1, other_min_h)  # 최소 1행어
                can_reduce = other_current - other_min_h
                if can_reduce > 0:
                    reduce = min(can_reduce, can_add)
                    templates_d0[r][other_g] -= reduce
                    templates_d0[r][g] = templates_d0[r].get(g, 0) + reduce
                    can_add -= reduce
                    hangers_to_add -= reduce

    # D+1 부족분 선반영 (용량 제한 적용)
    # 회전별 지그그룹별 사용량 계산
    d0_rotation_used = [{} for _ in range(10)]
    for r in range(10):
        for x in items:
            g = x['grp']
            if g not in d0_rotation_used[r]:
                d0_rotation_used[r][g] = 0
            d0_rotation_used[r][g] += x['prod'][r]

    for x in items:
        d0_end = x['stk']
        for r in range(10):
            d0_end = d0_end - x['d0'][r] + x['prod'][r]

        running = d0_end
        d1_deficit = 0
        for r in range(10):
            running = running - x['d1'][r]
            if running < 0:
                d1_deficit = max(d1_deficit, -running)

        if d1_deficit > 0:
            g = x['grp']
            remaining_deficit = d1_deficit
            # 뒤에서부터 여유 용량 내에서 추가
            for pr in range(9, -1, -1):
                if remaining_deficit <= 0:
                    break
                tmpl_h = templates_d0[pr].get(g, 0)
                rot_cap = tmpl_h * JIGS_PER_HANGER * JIG_INVENTORY.get(g, {'pcs': 1})['pcs']
                used = d0_rotation_used[pr].get(g, 0)
                available = max(0, rot_cap - used)
                if available > 0:
                    add = min(remaining_deficit, available)
                    x['prod'][pr] += add
                    d0_rotation_used[pr][g] = used + add
                    remaining_deficit -= add

    # =============================================
    # D+1 1-2회전 부족 그룹 내 재분배 (핵심 제약조건)
    # - 같은 그룹 내에서 과잉 → 부족으로 이동
    # =============================================
    def calc_d1_morning_shortage_final(x):
        """D0 생산만으로 D+1 1-2회전까지 부족 수량"""
        d0_end = x['stk']
        for r in range(10):
            d0_end = d0_end - x['d0'][r] + x['prod'][r]
        stk = d0_end
        max_shortage = 0
        for r in range(2):  # D+1-1, D+1-2만
            stk -= x.get('d1', [0]*10)[r]
            if stk < 0:
                max_shortage = max(max_shortage, -stk)
        return max_shortage

    def calc_d1_morning_excess_final(x):
        """D0 생산으로 D+1 1-2회전까지 여유 수량"""
        d0_end = x['stk']
        for r in range(10):
            d0_end = d0_end - x['d0'][r] + x['prod'][r]
        stk = d0_end
        min_stk = d0_end
        for r in range(2):  # D+1-1, D+1-2만
            stk -= x.get('d1', [0]*10)[r]
            min_stk = min(min_stk, stk)
        return max(0, min_stk)

    # D0 기말재고 계산 (빈행어 차감 전)
    for x in items:
        stk = x['stk']
        for r in range(10):
            stk = stk - x['d0'][r] + x['prod'][r]
        x['cur'] = stk

    # D+1 선반영 후 color_detail 및 컬러교환 재계산
    # (schedule_d0_optimized 이후 추가 생산이 발생했으므로)
    from collections import defaultdict as dd
    color_detail_d0 = []
    for r in range(10):
        detail = dd(int)
        for x in items:
            if x['prod'][r] > 0:
                detail[x['clr']] += x['prod'][r]
        color_detail_d0.append(dict(detail))

    # 컬러교환 재계산
    SPECIAL_COLORS = {'MGG', 'T4M', 'UMA', 'ZRM', 'ISM', 'MRM'}

    def get_cc_cost_local(from_clr, to_clr):
        return 15 if from_clr in SPECIAL_COLORS else 1

    def get_grp_colors_for_cc(grp, rot):
        clr_prod = dd(int)
        for x in items:
            if x['grp'] == grp and x['prod'][rot] > 0:
                clr_prod[x['clr']] += x['prod'][rot]
        if clr_prod:
            return sorted(clr_prod.keys(), key=lambda c: (-clr_prod[c], c))
        return []

    cc_count_d0 = 0
    cc_hangers_d0 = 0
    cc_count_per_rot_d0 = [0] * 10
    cc_hangers_per_rot_d0 = [0] * 10
    prev_clr = None

    for r in range(10):
        colors_ord = []
        order = jig_orders_d0[r]
        for g in order:
            if g in templates_d0[r] and templates_d0[r][g] > 0:
                grp_colors = get_grp_colors_for_cc(g, r)
                colors_ord.extend(grp_colors)

        # 회전간 컬러교환
        if prev_clr and colors_ord and colors_ord[0] != prev_clr:
            cc_count_per_rot_d0[r] += 1
            cc_hangers_per_rot_d0[r] += get_cc_cost_local(prev_clr, colors_ord[0])

        # 회전내 컬러교환
        for i in range(1, len(colors_ord)):
            if colors_ord[i] != colors_ord[i-1]:
                cc_count_per_rot_d0[r] += 1
                cc_hangers_per_rot_d0[r] += get_cc_cost_local(colors_ord[i-1], colors_ord[i])

        cc_count_d0 += cc_count_per_rot_d0[r]
        cc_hangers_d0 += cc_hangers_per_rot_d0[r]

        if colors_ord:
            prev_clr = colors_ord[-1]

    # D0→D+1 전환 지그교체
    d0_last_tmpl = templates_d0[9]
    d0_last_pos = order_to_positions(d0_last_tmpl, d0_last_order)

    # D+1 스케줄링
    (templates_d1, rotation_color_d1, jig_changes_d1, cc_count_d1, cc_hangers_d1,
     cc_count_per_rot_d1, cc_hangers_per_rot_d1,
     color_detail_d1, jig_orders_d1, d1_last_order, d1_last_color) = schedule_day_v2(
        items, 'd1', 'd1', 'prod1', 'cur',
        prev_template=d0_last_tmpl, prev_order=d0_last_order, prev_color=d0_last_color)

    d1_first_pos = order_to_positions(templates_d1[0], jig_orders_d1[0])
    d0_to_d1_jig = calc_position_changes(d0_last_pos, d1_first_pos)

    # D+1 기말재고
    for x in items:
        stk = x['cur']
        for r in range(10):
            stk = stk - x['d1'][r] + x['prod1'][r]
        x['cur1'] = stk

    # D+2 스케줄링
    d1_last_tmpl = templates_d1[9]
    d1_last_pos = order_to_positions(d1_last_tmpl, d1_last_order)

    (templates_d2, rotation_color_d2, jig_changes_d2, cc_count_d2, cc_hangers_d2,
     cc_count_per_rot_d2, cc_hangers_per_rot_d2,
     color_detail_d2, jig_orders_d2, d2_last_order, d2_last_color) = schedule_day_v2(
        items, 'd2', 'd2', 'prod2', 'cur1',
        prev_template=d1_last_tmpl, prev_order=d1_last_order, prev_color=d1_last_color)

    d2_first_pos = order_to_positions(templates_d2[0], jig_orders_d2[0])
    d1_to_d2_jig = calc_position_changes(d1_last_pos, d2_first_pos)

    # D+2 기말재고
    for x in items:
        stk = x['cur1']
        for r in range(10):
            stk = stk - x['d2'][r] + x['prod2'][r]
        x['cur2'] = stk

    # =============================================
    # 빈행어 손실 반영: 컬러교환으로 인한 생산량 차감
    # - 컬러교환 시 빈행어만큼 생산 불가능 (무조건 발생)
    # - 각 회전에서 빈행어 수만큼 생산량 감소
    # =============================================
    for r in range(10):
        empty_hangers = cc_hangers_per_rot_d0[r]
        if empty_hangers <= 0:
            continue

        # 이 회전에서 생산 중인 아이템들 (생산량 많은 것부터 차감)
        rot_items = [(x, x['prod'][r]) for x in items if x['prod'][r] > 0]
        rot_items.sort(key=lambda t: -t[1])

        remaining_loss = empty_hangers
        for x, prod in rot_items:
            if remaining_loss <= 0:
                break
            # 최소 0까지 차감 가능
            reduce = min(remaining_loss, prod)
            if reduce > 0:
                x['prod'][r] -= reduce
                remaining_loss -= reduce

    # =============================================
    # D+1 1-2회전 부족 그룹 내 재분배 (빈행어 차감 후)
    # - D+1-3~5는 D+1 생산으로 커버 가능, D+1-1~2만 D0 책임
    # =============================================
    def calc_d1_morning_shortage_post(x):
        """빈행어 차감 후 D+1 1-2회전까지 부족 수량"""
        d0_end = x['stk']
        for r in range(10):
            d0_end = d0_end - x['d0'][r] + x['prod'][r]
        stk = d0_end
        max_shortage = 0
        for r in range(2):  # D+1-1, D+1-2만
            stk -= x.get('d1', [0]*10)[r]
            if stk < 0:
                max_shortage = max(max_shortage, -stk)
        return max_shortage

    def calc_d1_morning_excess_post(x):
        """빈행어 차감 후 D+1 1-2회전까지 여유 수량"""
        d0_end = x['stk']
        for r in range(10):
            d0_end = d0_end - x['d0'][r] + x['prod'][r]
        stk = d0_end
        min_stk = d0_end
        for r in range(2):  # D+1-1, D+1-2만
            stk -= x.get('d1', [0]*10)[r]
            min_stk = min(min_stk, stk)
        return max(0, min_stk)

    # 그룹별로 재분배
    for g in set(x['grp'] for x in items):
        grp_items = [x for x in items if x['grp'] == g]

        for iteration in range(50):
            shortage_items_local = [(x, calc_d1_morning_shortage_post(x)) for x in grp_items]
            shortage_items_local = [(x, s) for x, s in shortage_items_local if s > 0]
            if not shortage_items_local:
                break

            excess_items_local = [(x, calc_d1_morning_excess_post(x)) for x in grp_items]
            excess_items_local = [(x, e) for x, e in excess_items_local if e > 0 and sum(x['prod']) > 0]
            if not excess_items_local:
                break

            shortage_items_local.sort(key=lambda t: -t[1])
            excess_items_local.sort(key=lambda t: -t[1])

            shortage_x, shortage_amt = shortage_items_local[0]
            excess_x, excess_amt = excess_items_local[0]

            if shortage_x == excess_x:
                break

            moved = 0
            for r in range(9, -1, -1):
                if moved >= shortage_amt or moved >= excess_amt:
                    break
                if excess_x['prod'][r] > 0:
                    transfer = min(shortage_amt - moved, excess_amt - moved, excess_x['prod'][r])
                    if transfer > 0:
                        excess_x['prod'][r] -= transfer
                        shortage_x['prod'][r] += transfer
                        moved += transfer

            if moved == 0:
                break

    # =============================================
    # 2회전 리드타임 제약 적용
    # - R회전 생산 → R+2회전부터 사용 가능
    # - R회전 부족 → R-2회전까지 생산 필요
    # =============================================
    LEAD_TIME_POST = 2

    def calc_lead_time_shortage(x):
        """리드타임 적용 시 각 회전별 부족 계산 (D0 + D+1 1-2회전만)"""
        stk = x['stk']
        shortage_by_rot = []

        for r in range(10):  # D0
            stk -= x['d0'][r]
            # R-2회전 생산이 사용 가능
            if r >= LEAD_TIME_POST:
                stk += x['prod'][r - LEAD_TIME_POST]
            if stk < 0:
                shortage_by_rot.append((r, -stk))

        # D+1 1-2회전만 (D+1-3~5는 D+1 생산으로 커버)
        for r in range(2):
            stk -= x.get('d1', [0]*10)[r]
            d0_rot = 8 + r  # D0-9, D0-10 생산이 D+1-1, D+1-2에 사용 가능
            if d0_rot < 10:
                stk += x['prod'][d0_rot]
            if stk < 0:
                shortage_by_rot.append((10 + r, -stk))

        return shortage_by_rot

    # 리드타임 부족 아이템 찾기 및 생산 앞당기기
    for g in set(x['grp'] for x in items):
        grp_items = [x for x in items if x['grp'] == g]

        for iteration in range(30):
            # 리드타임 부족 있는 아이템 찾기
            shortage_item = None
            shortage_rot = None
            shortage_amt = 0
            for x in grp_items:
                shortages = calc_lead_time_shortage(x)
                if shortages:
                    shortage_item = x
                    shortage_rot, shortage_amt = shortages[0]
                    break

            if not shortage_item:
                break

            # 같은 그룹 내 다른 아이템에서 앞 회전 생산을 빌려옴
            # 또는 같은 아이템 내에서 뒤 회전 생산을 앞으로 이동
            moved = False

            # 필요한 생산 회전 (리드타임 고려, 최대 D0-10까지)
            need_by_rot = min(9, max(0, shortage_rot - LEAD_TIME_POST))

            # 1) 같은 그룹 다른 아이템에서 앞 회전 생산 빌려오기
            for other_x in grp_items:
                if moved:
                    break
                if other_x == shortage_item:
                    continue
                # 다른 아이템의 여유 계산
                other_excess = calc_d1_morning_excess_post(other_x)
                if other_excess <= 0:
                    continue
                # 앞 회전(need_by_rot 이하)에서 생산 있는지
                for src_r in range(min(10, need_by_rot + 1)):
                    if other_x['prod'][src_r] > 0:
                        transfer = min(shortage_amt, other_excess, other_x['prod'][src_r])
                        if transfer > 0:
                            other_x['prod'][src_r] -= transfer
                            shortage_item['prod'][src_r] += transfer
                            moved = True
                            break

            # 2) 같은 아이템 내에서 뒤 회전 생산을 앞으로 이동 (템플릿 용량 허용 시)
            if not moved:
                for src_r in range(9, need_by_rot, -1):
                    if shortage_item['prod'][src_r] > 0:
                        # src_r 생산을 need_by_rot으로 이동
                        transfer = min(shortage_amt, shortage_item['prod'][src_r])
                        if transfer > 0:
                            shortage_item['prod'][src_r] -= transfer
                            shortage_item['prod'][need_by_rot] += transfer
                            moved = True
                            break

            if not moved:
                break

    # D0 기말재고 재계산
    for x in items:
        stk = x['stk']
        for r in range(10):
            stk = stk - x['d0'][r] + x['prod'][r]
        x['cur'] = stk

    # color_detail 재계산
    color_detail_d0 = []
    for r in range(10):
        detail = dd(int)
        for x in items:
            if x['prod'][r] > 0:
                detail[x['clr']] += x['prod'][r]
        color_detail_d0.append(dict(detail))

    return {
        'd0': {
            'templates': templates_d0,
            'colors': rotation_color_d0,
            'jig_changes': jig_changes_d0,
            'cc_count': cc_count_d0,
            'cc_hangers': cc_hangers_d0,
            'cc_count_per_rotation': cc_count_per_rot_d0,
            'cc_hangers_per_rotation': cc_hangers_per_rot_d0,
            'color_detail': color_detail_d0,
            'jig_orders': jig_orders_d0,
        },
        'd1': {
            'templates': templates_d1,
            'colors': rotation_color_d1,
            'jig_changes': jig_changes_d1,
            'cc_count': cc_count_d1,
            'cc_hangers': cc_hangers_d1,
            'cc_count_per_rotation': cc_count_per_rot_d1,
            'cc_hangers_per_rotation': cc_hangers_per_rot_d1,
            'color_detail': color_detail_d1,
            'jig_orders': jig_orders_d1,
            'start_jig_change': d0_to_d1_jig
        },
        'd2': {
            'templates': templates_d2,
            'colors': rotation_color_d2,
            'jig_changes': jig_changes_d2,
            'cc_count': cc_count_d2,
            'cc_hangers': cc_hangers_d2,
            'cc_count_per_rotation': cc_count_per_rot_d2,
            'cc_hangers_per_rotation': cc_hangers_per_rot_d2,
            'color_detail': color_detail_d2,
            'jig_orders': jig_orders_d2,
            'start_jig_change': d1_to_d2_jig
        },
    }


# ============================================
# 리포트 생성 함수
# ============================================

def get_rotation_items_detail(items, rotation, prod_key, templates, jig_orders):
    """회전별 생산 아이템 상세 - 컬러별로 묶어서 표시"""
    tmpl = templates[rotation]
    order = jig_orders[rotation] if jig_orders and jig_orders[rotation] else sorted(tmpl.keys())

    result = []
    for g in order:
        if g not in tmpl or tmpl[g] == 0:
            continue
        grp_items = [(x, x[prod_key][rotation]) for x in items
                     if x['grp'] == g and x[prod_key][rotation] > 0]

        # 컬러별 생산량 합계 계산
        color_totals = {}
        for x, prod in grp_items:
            clr = x['clr']
            color_totals[clr] = color_totals.get(clr, 0) + prod

        # 컬러 순서: 생산량 많은 순, 같으면 알파벳 순
        color_order = sorted(color_totals.keys(), key=lambda c: (-color_totals[c], c))
        color_rank = {c: i for i, c in enumerate(color_order)}

        # 아이템 정렬: 컬러 순서 → 생산량 순
        grp_items.sort(key=lambda x: (color_rank.get(x[0]['clr'], 999), -x[1]))

        if grp_items:
            for x, prod in grp_items:
                ct = x['ct'].replace('\n', ' ').strip()
                it = x['it'].replace('\n', ' ').strip()
                det = x['det'].replace('\n', ' ').strip() if x.get('det') else '-'
                result.append((ct, it, det, x['clr'], prod, g))
        else:
            # 생산 0이어도 그룹 표시 (행어바와 일치)
            result.append(('-', '-', '-', '-', 0, g))
    return result


def format_rotation_items_html(items, rotation, prod_key, templates, jig_orders):
    """회전별 생산 아이템을 HTML 박스 형태로 포맷"""
    details = get_rotation_items_detail(items, rotation, prod_key, templates, jig_orders)
    if not details:
        return "<span style='color:#999;'>-</span>"

    grp_colors = {
        'A': '#E3F2FD', 'B': '#E8F5E9', 'B2': '#C8E6C9', 'C': '#FFF3E0', 'D': '#FCE4EC',
        'E': '#F3E5F5', 'F': '#E0F7FA', 'G': '#FFF8E1', 'H': '#EFEBE9', 'I': '#ECEFF1'
    }
    grp_border = {
        'A': '#1976D2', 'B': '#388E3C', 'B2': '#2E7D32', 'C': '#F57C00', 'D': '#C2185B',
        'E': '#7B1FA2', 'F': '#0097A7', 'G': '#FFA000', 'H': '#5D4037', 'I': '#455A64'
    }

    result_parts = []
    current_grp = None
    grp_items = []

    tmpl = templates[rotation]

    for ct, it, det, clr, prod, g in details:
        if g != current_grp:
            if grp_items and current_grp:
                bg = grp_colors.get(current_grp, '#F5F5F5')
                border = grp_border.get(current_grp, '#9E9E9E')
                h_count = tmpl.get(current_grp, 0)
                items_html = "<br>".join([f"&nbsp;{x}" for x in grp_items])
                result_parts.append(
                    f'<div style="display:inline-block;vertical-align:top;margin:2px;padding:4px 8px;'
                    f'background:{bg};border:2px solid {border};border-radius:6px;min-width:150px;">'
                    f'<div style="font-weight:bold;color:{border};border-bottom:1px solid {border};margin-bottom:3px;padding-bottom:2px;">'
                    f'{current_grp} ({h_count}H)</div>'
                    f'<div style="font-size:0.7em;line-height:1.5;">{items_html}</div>'
                    f'</div>'
                )
            current_grp = g
            grp_items = []

        if prod > 0:
            det_str = f" {det}" if det and det != '-' else ""
            grp_items.append(f"<b>{ct}</b> {it}{det_str} <span style='color:#1565C0;'>{clr}</span> <span style='color:#D32F2F;font-weight:bold;'>{prod}</span>")
        else:
            # 생산 0인 경우
            grp_items.append(f"<span style='color:#999;'>생산없음</span>")

    if grp_items and current_grp:
        bg = grp_colors.get(current_grp, '#F5F5F5')
        border = grp_border.get(current_grp, '#9E9E9E')
        h_count = tmpl.get(current_grp, 0)
        items_html = "<br>".join([f"&nbsp;{x}" for x in grp_items])
        result_parts.append(
            f'<div style="display:inline-block;vertical-align:top;margin:2px;padding:4px 8px;'
            f'background:{bg};border:2px solid {border};border-radius:6px;min-width:150px;">'
            f'<div style="font-weight:bold;color:{border};border-bottom:1px solid {border};margin-bottom:3px;padding-bottom:2px;">'
            f'{current_grp} ({h_count}H)</div>'
            f'<div style="font-size:0.7em;line-height:1.5;">{items_html}</div>'
            f'</div>'
        )

    return "".join(result_parts)


def format_hanger_positions_html(templates, jig_orders, rotation, prev_positions=None,
                                  items=None, prod_key='prod', prev_grp_colors=None):
    """140개 행어 위치를 시각적으로 표시

    - 지그 교체: 빨간 왼쪽 테두리 (굵은 선)
    - 컬러 교체: 빨간 점선 왼쪽 테두리
    """
    tmpl = templates[rotation]
    order = jig_orders[rotation] if jig_orders and jig_orders[rotation] else sorted(tmpl.keys())

    curr_positions = order_to_positions(tmpl, order)

    grp_colors_map = {
        'A': '#1976D2', 'B': '#388E3C', 'B2': '#2E7D32', 'C': '#F57C00', 'D': '#C2185B',
        'E': '#7B1FA2', 'F': '#0097A7', 'G': '#FFA000', 'H': '#5D4037', 'I': '#455A64'
    }

    # 그룹별 주요 컬러 계산
    curr_grp_colors = {}
    if items:
        for g in order:
            color_prod = defaultdict(int)
            for x in items:
                if x.get('grp') == g:
                    prod = x.get(prod_key, [0]*10)
                    if rotation < len(prod) and prod[rotation] > 0:
                        color_prod[x['clr']] += prod[rotation]
            if color_prod:
                curr_grp_colors[g] = max(color_prod.keys(), key=lambda c: color_prod[c])

    segments = []
    if curr_positions:
        start = 0
        current_grp = curr_positions[0]
        for i in range(1, len(curr_positions)):
            if curr_positions[i] != current_grp:
                segments.append({
                    'grp': current_grp,
                    'start': start,
                    'end': i - 1,
                    'count': i - start
                })
                start = i
                current_grp = curr_positions[i]
        segments.append({
            'grp': current_grp,
            'start': start,
            'end': len(curr_positions) - 1,
            'count': len(curr_positions) - start
        })

    # 지그 교체 위치
    jig_change_positions = set()
    if prev_positions:
        for i in range(HANGERS):
            if i < len(prev_positions) and i < len(curr_positions):
                if prev_positions[i] != curr_positions[i]:
                    jig_change_positions.add(i)

    html_parts = []
    html_parts.append('<div style="display:flex;align-items:center;gap:10px;margin:5px 0;">')
    html_parts.append('<span style="font-size:0.75em;color:#666;min-width:60px;">행어위치:</span>')
    html_parts.append('<div style="display:flex;flex-wrap:nowrap;border:1px solid #999;border-radius:4px;overflow:hidden;flex:1;">')

    prev_seg_color = None
    color_changes = 0
    for idx, seg in enumerate(segments):
        grp = seg['grp']
        count = seg['count']
        start = seg['start']
        end = seg['end']

        if grp is None:
            color = '#E0E0E0'
            label = '-'
            seg_paint_color = None
        else:
            color = grp_colors_map.get(grp, '#9E9E9E')
            label = grp
            seg_paint_color = curr_grp_colors.get(grp)

        # 지그 교체 여부
        has_jig_change = any(pos in jig_change_positions for pos in range(start, end + 1))

        # 컬러 교체 여부 (이전 세그먼트와 비교)
        has_color_change = False
        if idx > 0 and prev_seg_color and seg_paint_color:
            if prev_seg_color != seg_paint_color:
                has_color_change = True
                color_changes += 1

        # 스타일 결정
        if has_jig_change and start > 0:
            border_style = 'border-left:3px solid #F44336;'  # 지그 교체: 굵은 빨간선
        elif has_color_change:
            border_style = 'border-left:3px dashed #F44336;'  # 컬러 교체: 빨간 점선
        else:
            border_style = ''

        width_pct = (count / HANGERS) * 100

        # 컬러 정보 표시
        color_label = f' [{seg_paint_color[:3]}]' if seg_paint_color else ''

        html_parts.append(
            f'<div style="width:{width_pct:.1f}%;background:{color};color:white;text-align:center;'
            f'font-size:0.7em;padding:2px 0;min-width:15px;{border_style}" '
            f'title="위치 {start+1}-{end+1} ({count}행어) {seg_paint_color or ""}">'
            f'{label}<span style="font-size:0.8em;opacity:0.8;">({count})</span></div>'
        )

        prev_seg_color = seg_paint_color

    html_parts.append('</div>')

    # 교체 정보 표시
    info_parts = []
    if jig_change_positions:
        info_parts.append(f'<span style="color:#D32F2F;">지그:{len(jig_change_positions)}</span>')
    if color_changes > 0:
        info_parts.append(f'<span style="color:#D32F2F;">컬러:{color_changes}</span>')
    if info_parts:
        html_parts.append(f'<span style="font-size:0.7em;margin-left:5px;">{" ".join(info_parts)}</span>')

    html_parts.append('</div>')

    return ''.join(html_parts), curr_positions, curr_grp_colors


def generate_html_report(items, schedule_result):
    today = datetime.now()
    tomorrow = today + timedelta(days=1)
    day_after = today + timedelta(days=2)

    d0 = schedule_result['d0']
    d1 = schedule_result['d1']
    d2 = schedule_result['d2']

    templates = d0['templates']
    rotation_color = d0['colors']
    jig_changes = d0['jig_changes']
    cc_count = d0['cc_count']  # 컬러교환 횟수
    cc_hangers = d0['cc_hangers']  # 빈행어 손실
    odd_jig_loss = d0.get('odd_jig_loss', 0)  # 홀수 생산 손실
    jig_orders_d0 = d0.get('jig_orders', [None]*10)
    jig_orders_d1 = d1.get('jig_orders', [None]*10)
    jig_orders_d2 = d2.get('jig_orders', [None]*10)

    day_jig = sum(jig_changes[:5])
    night_jig = sum(jig_changes[5:])
    total_prod_d0 = sum(sum(x['prod']) for x in items)
    total_prod_d1 = sum(sum(x['prod1']) for x in items)
    total_prod_d2 = sum(sum(x['prod2']) for x in items)

    html = f'''<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>D0 생산계획 리포트 - {today.strftime("%Y-%m-%d")}</title>
    <style>
        * {{ box-sizing: border-box; }}
        body {{
            font-family: 'Malgun Gothic', sans-serif;
            margin: 20px;
            background: #f5f5f5;
        }}
        h1 {{ color: #333; border-bottom: 3px solid #2196F3; padding-bottom: 10px; }}
        h2 {{ color: #1976D2; margin-top: 30px; }}
        h3 {{ color: #424242; }}
        .container {{ max-width: 1800px; margin: 0 auto; }}
        .card {{
            background: white;
            border-radius: 8px;
            padding: 20px;
            margin: 15px 0;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .params-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 15px;
        }}
        .param-item {{
            background: #E3F2FD;
            padding: 12px;
            border-radius: 6px;
            border-left: 4px solid #2196F3;
        }}
        .param-label {{ font-weight: bold; color: #1565C0; }}
        .param-value {{ font-size: 1.2em; color: #333; margin-top: 5px; }}
        .jig-table {{
            width: 100%;
            border-collapse: collapse;
            margin: 10px 0;
        }}
        .jig-table th, .jig-table td {{
            border: 1px solid #ddd;
            padding: 8px;
            text-align: center;
        }}
        .jig-table th {{
            background: #2196F3;
            color: white;
        }}
        .jig-table tr:nth-child(even) {{ background: #f9f9f9; }}
        .summary-box {{
            display: flex;
            gap: 20px;
            flex-wrap: wrap;
        }}
        .summary-item {{
            flex: 1;
            min-width: 150px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
        }}
        .summary-item.warning {{
            background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        }}
        .summary-item.success {{
            background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
        }}
        .summary-item.over-budget {{
            background: linear-gradient(135deg, #ff416c 0%, #ff4b2b 100%);
        }}
        .summary-number {{ font-size: 2em; font-weight: bold; }}
        .summary-label {{ font-size: 0.9em; opacity: 0.9; }}
        .main-table-container {{
            overflow-x: auto;
            max-height: 800px;
            overflow-y: auto;
        }}
        .main-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 0.85em;
        }}
        .main-table th {{
            background: #37474F;
            color: white;
            padding: 8px 4px;
            position: sticky;
            top: 0;
            z-index: 10;
        }}
        .main-table td {{
            border: 1px solid #ddd;
            padding: 4px;
            text-align: center;
        }}
        .main-table tr:nth-child(even) {{ background: #fafafa; }}
        .main-table tr:hover {{ background: #e3f2fd; }}
        .shortage {{
            background: #FFCDD2 !important;
            color: #B71C1C;
            font-weight: bold;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>D0 생산계획 리포트 v9.2</h1>
        <p>생성일시: {today.strftime("%Y-%m-%d %H:%M:%S")}</p>

        <div class="card">
            <h2>시스템 변수</h2>
            <div class="params-grid">
                <div class="param-item">
                    <div class="param-label">총 행어 수</div>
                    <div class="param-value">{HANGERS}개</div>
                </div>
                <div class="param-item">
                    <div class="param-label">행어당 지그 수</div>
                    <div class="param-value">{JIGS_PER_HANGER}개</div>
                </div>
                <div class="param-item">
                    <div class="param-label">일일 회전 수</div>
                    <div class="param-value">{ROTATIONS_PER_DAY}회전</div>
                </div>
                <div class="param-item">
                    <div class="param-label">주간 행어교체 예산</div>
                    <div class="param-value">{HANGER_BUDGET_DAY}개</div>
                </div>
                <div class="param-item">
                    <div class="param-label">야간 행어교체 예산</div>
                    <div class="param-value">{HANGER_BUDGET_NIGHT}개</div>
                </div>
                <div class="param-item">
                    <div class="param-label">컬러교환 빈행어</div>
                    <div class="param-value">특수 15행어 / 일반 1행어</div>
                </div>
            </div>
        </div>

        <div class="card">
            <h2>지그 그룹 정보</h2>
            <table class="jig-table">
                <tr>
                    <th>그룹</th>
                    <th>명칭</th>
                    <th>최대 지그수</th>
                    <th>최대 행어</th>
                    <th>PCS/지그</th>
                    <th>최대 용량/회전</th>
                </tr>'''

    for g in sorted(JIG_INVENTORY.keys()):
        info = JIG_INVENTORY[g]
        max_h = info['max_jigs'] // JIGS_PER_HANGER
        max_cap = max_h * JIGS_PER_HANGER * info['pcs']
        html += f'''
                <tr>
                    <td><strong>{g}</strong></td>
                    <td>{info['name']}</td>
                    <td>{info['max_jigs']}</td>
                    <td>{max_h}</td>
                    <td>{info['pcs']}</td>
                    <td>{max_cap}</td>
                </tr>'''

    day_jig_class = 'over-budget' if day_jig > HANGER_BUDGET_DAY else 'success'
    night_jig_class = 'over-budget' if night_jig > HANGER_BUDGET_NIGHT else 'success'

    html += f'''
            </table>
        </div>

        <div class="card">
            <h2>실행 결과 요약</h2>
            <div class="summary-box">
                <div class="summary-item success">
                    <div class="summary-number">{total_prod_d0:,}</div>
                    <div class="summary-label">D0 생산량</div>
                </div>
                <div class="summary-item success">
                    <div class="summary-number">{total_prod_d1:,}</div>
                    <div class="summary-label">D+1 생산량</div>
                </div>
                <div class="summary-item success">
                    <div class="summary-number">{total_prod_d2:,}</div>
                    <div class="summary-label">D+2 생산량</div>
                </div>
                <div class="summary-item">
                    <div class="summary-number">{cc_count}</div>
                    <div class="summary-label">D0 컬러교환</div>
                </div>
                <div class="summary-item warning">
                    <div class="summary-number">{cc_hangers}</div>
                    <div class="summary-label">D0 빈행어 손실</div>
                </div>
                <div class="summary-item warning">
                    <div class="summary-number">{odd_jig_loss}</div>
                    <div class="summary-label">D0 홀수 손실</div>
                </div>
                <div class="summary-item warning">
                    <div class="summary-number">{cc_hangers * 2 + odd_jig_loss}</div>
                    <div class="summary-label">D0 총 손실</div>
                </div>
                <div class="summary-item {day_jig_class}">
                    <div class="summary-number">{day_jig}/{HANGER_BUDGET_DAY}</div>
                    <div class="summary-label">D0 주간행어</div>
                </div>
                <div class="summary-item {night_jig_class}">
                    <div class="summary-number">{night_jig}/{HANGER_BUDGET_NIGHT}</div>
                    <div class="summary-label">D0 야간행어</div>
                </div>
            </div>
        </div>

        <div class="card">
            <h2>D0 생산 없을 때 재고부족 현황 (D+1 오전까지)</h2>
            <p style="color:#666;margin-bottom:15px;">D0에 생산하지 않으면 D+1 오전(5회전)까지 언제 재고부족이 발생하는지 보여줍니다.</p>
'''
    # 생산 없을 때 D+1 오전까지 재고부족 계산
    shortage_items = []
    rot_shortages = {}  # 회전별 그룹화

    for x in items:
        stk = x['stk']
        shortage_rot = -1
        first_shortage_amt = 0

        # D0 각 회전에서 부족 체크 (생산 0 가정)
        for r in range(10):
            stk = stk - x['d0'][r]
            if stk < 0 and shortage_rot < 0:
                shortage_rot = r + 1  # D0-1 ~ D0-10
                first_shortage_amt = -stk

        # D+1 오전 5회전 체크
        for r in range(5):
            stk = stk - x['d1'][r]
            if stk < 0 and shortage_rot < 0:
                shortage_rot = 10 + r + 1  # D+1-1 ~ D+1-5 (11~15)
                first_shortage_amt = -stk

        # 최종 부족량 (D+1 오전 끝 시점)
        final_shortage = -stk if stk < 0 else 0

        if shortage_rot > 0:
            rot_key = shortage_rot
            if rot_key not in rot_shortages:
                rot_shortages[rot_key] = []
            rot_shortages[rot_key].append({
                'ct': x['ct'],
                'it': x['it'],
                'det': x.get('det', '-'),
                'clr': x['clr'],
                'grp': x['grp'],
                'stk': x['stk'],
                'd0t': sum(x['d0']),
                'd1_am': sum(x['d1'][:5]),
                'rot': shortage_rot,
                'first_amt': first_shortage_amt,
                'final_amt': final_shortage
            })
            shortage_items.append(rot_shortages[rot_key][-1])

    # 시간대별 요약 테이블
    html += '''
            <h3 style="margin-top:0;">시간대별 부족 요약</h3>
            <table class="jig-table" style="font-size:0.9em;max-width:800px;">
                <tr>
                    <th>시간대</th>
                    <th>부족 시작 아이템 수</th>
                    <th>부족량 합계</th>
                    <th>비고</th>
                </tr>
'''
    # D0 주간 (1-5회전)
    d0_day_items = sum(len(rot_shortages.get(r, [])) for r in range(1, 6))
    d0_day_amt = sum(s['first_amt'] for r in range(1, 6) for s in rot_shortages.get(r, []))
    # D0 야간 (6-10회전)
    d0_night_items = sum(len(rot_shortages.get(r, [])) for r in range(6, 11))
    d0_night_amt = sum(s['first_amt'] for r in range(6, 11) for s in rot_shortages.get(r, []))
    # D+1 오전 (1-5회전 = 11-15)
    d1_am_items = sum(len(rot_shortages.get(r, [])) for r in range(11, 16))
    d1_am_amt = sum(s['first_amt'] for r in range(11, 16) for s in rot_shortages.get(r, []))

    html += f'''
                <tr>
                    <td style="background:#E3F2FD;font-weight:bold;">D0 주간 (1-5회전)</td>
                    <td>{d0_day_items}건</td>
                    <td style="color:#D32F2F;font-weight:bold;">{d0_day_amt}개</td>
                    <td>{'없음' if d0_day_items == 0 else '긴급 생산 필요'}</td>
                </tr>
                <tr>
                    <td style="background:#E8EAF6;font-weight:bold;">D0 야간 (6-10회전)</td>
                    <td>{d0_night_items}건</td>
                    <td style="color:#D32F2F;font-weight:bold;">{d0_night_amt}개</td>
                    <td>{'없음' if d0_night_items == 0 else '주간 선생산 필요'}</td>
                </tr>
                <tr>
                    <td style="background:#FFF3E0;font-weight:bold;">D+1 오전 (1-5회전)</td>
                    <td>{d1_am_items}건</td>
                    <td style="color:#E65100;font-weight:bold;">{d1_am_amt}개</td>
                    <td>{'없음' if d1_am_items == 0 else 'D0 생산으로 커버'}</td>
                </tr>
                <tr style="background:#FFEBEE;font-weight:bold;">
                    <td>합계</td>
                    <td>{len(shortage_items)}건</td>
                    <td style="color:#D32F2F;">{d0_day_amt + d0_night_amt + d1_am_amt}개</td>
                    <td></td>
                </tr>
            </table>

            <h3 style="margin-top:25px;">D0 생산 없을 때 3일 재고 전망 (부족 아이템)</h3>
            <p style="color:#666;font-size:0.85em;">D0 생산이 전혀 없다고 가정할 때 각 회전별 재고 흐름 (부족 발생 아이템만 표시)</p>
        </div>
        <div class="card" style="overflow-x:auto;">
        <table class="jig-table" style="font-size:0.75em;">
        <thead>
            <tr>
                <th rowspan="2">차종</th>
                <th rowspan="2">아이템</th>
                <th rowspan="2">세부</th>
                <th rowspan="2">컬러</th>
                <th rowspan="2">지그</th>
                <th rowspan="2">기초</th>'''

    # D0 수요/재고 헤더
    for i in range(10):
        html += f'<th colspan="2" style="background:#1976D2;">D0-{i+1}</th>'
    # D+1 오전 5회전만
    for i in range(5):
        html += f'<th colspan="2" style="background:#388E3C;">D+1-{i+1}</th>'
    html += '<th rowspan="2">최종</th>'

    html += '</tr><tr>'
    # D0 수요/재고
    for i in range(10):
        html += '<th style="background:#1976D2;font-size:0.8em;">수</th>'
        html += '<th style="background:#1976D2;font-size:0.8em;">재</th>'
    # D+1 오전
    for i in range(5):
        html += '<th style="background:#388E3C;font-size:0.8em;">수</th>'
        html += '<th style="background:#388E3C;font-size:0.8em;">재</th>'
    html += '</tr></thead><tbody>'

    # 부족 아이템 정렬 (부족 심한 순)
    def get_min_stock_no_prod(x):
        stk = x['stk']
        min_stk = stk
        for i in range(10):
            stk = stk - x['d0'][i]
            min_stk = min(min_stk, stk)
        for i in range(5):
            stk = stk - x['d1'][i]
            min_stk = min(min_stk, stk)
        return min_stk

    items_with_shortage = [x for x in items if get_min_stock_no_prod(x) < 0]
    items_with_shortage.sort(key=get_min_stock_no_prod)

    if items_with_shortage:
        for x in items_with_shortage:
            ct = x['ct'].replace('\n', ' ')[:10]
            it = x['it'].replace('\n', ' ')[:8]
            det = (x['det'] or '-').replace('\n', ' ')[:6]
            clr = x['clr'][:5] if x['clr'] else '-'
            grp = x['grp'] or '-'

            html += f'<tr style="background:#FFF3E0;"><td>{ct}</td><td>{it}</td><td>{det}</td><td>{clr}</td><td>{grp}</td>'
            html += f'<td>{x["stk"]}</td>'

            # D0 (생산 없음)
            stk = x['stk']
            for i in range(10):
                dem = x['d0'][i]
                stk = stk - dem
                stk_style = 'color:#D32F2F;font-weight:bold;background:#FFCDD2;' if stk < 0 else ''
                html += f'<td>{dem}</td><td style="{stk_style}">{stk}</td>'

            # D+1 오전 5회전 (생산 없음)
            for i in range(5):
                dem = x['d1'][i]
                stk = stk - dem
                stk_style = 'color:#D32F2F;font-weight:bold;background:#FFCDD2;' if stk < 0 else ''
                html += f'<td>{dem}</td><td style="{stk_style}">{stk}</td>'

            final_style = 'color:#D32F2F;font-weight:bold;background:#FFCDD2;' if stk < 0 else 'font-weight:bold;'
            html += f'<td style="{final_style}">{stk}</td></tr>'
    else:
        html += '''<tr><td colspan="37" style="color:#4CAF50;font-weight:bold;padding:20px;">
            생산 없이도 D+1 오전까지 재고부족 없음</td></tr>'''

    html += '''</tbody></table>
        <p style="color:#888;font-size:0.85em;margin-top:10px;">
            * D0 생산이 전혀 없다고 가정할 때의 재고 흐름<br/>
            * 빨간색 셀은 해당 시점에서 재고 부족 발생
        </p>
        </div>
'''

    # 재고/수요/생산 합계
    init_stock = sum(x['stk'] for x in items)
    d0_demand = sum(x['d0t'] for x in items)
    d0_prod = total_prod_d0
    d0_end_stock = sum(x['cur'] for x in items)
    d1_demand = sum(x['d1t'] for x in items)
    d1_prod = total_prod_d1
    d1_end_stock = sum(x['cur1'] for x in items)
    d2_demand = sum(x['d2t'] for x in items)
    d2_prod = total_prod_d2
    d2_end_stock = sum(x['cur2'] for x in items)

    html += f'''
        <div class="card">
            <h2>재고/수요/생산 합계</h2>
            <table class="summary-table" style="width:100%;border-collapse:collapse;text-align:center;">
                <tr style="background:#37474F;color:white;">
                    <th style="padding:10px;border:1px solid #ccc;">구분</th>
                    <th style="padding:10px;border:1px solid #ccc;">기초재고</th>
                    <th style="padding:10px;border:1px solid #ccc;background:#1565C0;">D0 수요</th>
                    <th style="padding:10px;border:1px solid #ccc;background:#1565C0;">D0 생산</th>
                    <th style="padding:10px;border:1px solid #ccc;background:#1565C0;">D0 기말재고</th>
                    <th style="padding:10px;border:1px solid #ccc;background:#2E7D32;">D+1 수요</th>
                    <th style="padding:10px;border:1px solid #ccc;background:#2E7D32;">D+1 생산</th>
                    <th style="padding:10px;border:1px solid #ccc;background:#2E7D32;">D+1 기말재고</th>
                    <th style="padding:10px;border:1px solid #ccc;background:#E65100;">D+2 수요</th>
                    <th style="padding:10px;border:1px solid #ccc;background:#E65100;">D+2 생산</th>
                    <th style="padding:10px;border:1px solid #ccc;background:#E65100;">D+2 기말재고</th>
                </tr>
                <tr style="font-size:1.2em;font-weight:bold;">
                    <td style="padding:12px;border:1px solid #ccc;background:#ECEFF1;">합계</td>
                    <td style="padding:12px;border:1px solid #ccc;">{init_stock:,}</td>
                    <td style="padding:12px;border:1px solid #ccc;background:#E3F2FD;">{d0_demand:,}</td>
                    <td style="padding:12px;border:1px solid #ccc;background:#E3F2FD;color:#1565C0;">{d0_prod:,}</td>
                    <td style="padding:12px;border:1px solid #ccc;background:#E3F2FD;">{d0_end_stock:,}</td>
                    <td style="padding:12px;border:1px solid #ccc;background:#E8F5E9;">{d1_demand:,}</td>
                    <td style="padding:12px;border:1px solid #ccc;background:#E8F5E9;color:#2E7D32;">{d1_prod:,}</td>
                    <td style="padding:12px;border:1px solid #ccc;background:#E8F5E9;">{d1_end_stock:,}</td>
                    <td style="padding:12px;border:1px solid #ccc;background:#FFF3E0;">{d2_demand:,}</td>
                    <td style="padding:12px;border:1px solid #ccc;background:#FFF3E0;color:#E65100;">{d2_prod:,}</td>
                    <td style="padding:12px;border:1px solid #ccc;background:#FFF3E0;">{d2_end_stock:,}</td>
                </tr>
            </table>
        </div>
'''

    # 회전별 생산 상세
    templates_d1 = d1['templates']
    jig_changes_d1 = d1['jig_changes']
    templates_d2_box = d2['templates']
    jig_changes_d2_box = d2['jig_changes']

    html += '''
        <div class="card">
            <h2>회전별 생산 상세 (지그 순서)</h2>
            <p>각 회전에서 생산되는 아이템을 컨베이어 지그 순서대로 표시 (박스 = 지그그룹, 숫자 = 행어수)</p>
'''

    # 지그그룹 색상 범례
    grp_colors_legend = {
        'A': '#1976D2', 'B': '#388E3C', 'B2': '#2E7D32', 'C': '#F57C00', 'D': '#C2185B',
        'E': '#7B1FA2', 'F': '#0097A7', 'G': '#FFA000', 'H': '#5D4037', 'I': '#455A64'
    }
    legend_html = '<div style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:10px;font-size:0.75em;">'
    for g, color in grp_colors_legend.items():
        legend_html += f'<span style="background:{color};color:white;padding:2px 6px;border-radius:3px;">{g}</span>'
    legend_html += '<span style="border-left:3px solid #F44336;padding-left:5px;margin-left:10px;">= 지그교체</span>'
    legend_html += '<span style="border-left:3px dashed #F44336;padding-left:5px;margin-left:10px;">= 컬러교체</span></div>'

    # D0
    html += '<h3 style="margin-top:20px;color:#1565C0;">D0 생산계획</h3>'
    html += legend_html

    prev_positions_d0 = None
    prev_grp_colors_d0 = None
    for r in range(10):
        shift_name = '주간' if r < 5 else '야간'
        shift_bg = '#E3F2FD' if r < 5 else '#E8EAF6'
        detail_html = format_rotation_items_html(items, r, 'prod', templates, jig_orders_d0)
        hanger_html, curr_positions, curr_grp_colors = format_hanger_positions_html(
            templates, jig_orders_d0, r, prev_positions_d0,
            items=items, prod_key='prod', prev_grp_colors=prev_grp_colors_d0)
        prev_positions_d0 = curr_positions
        prev_grp_colors_d0 = curr_grp_colors

        html += f'''
            <div style="margin:8px 0;padding:10px;background:{shift_bg};border-radius:8px;">
                <div style="display:flex;align-items:center;gap:15px;margin-bottom:8px;">
                    <span style="font-weight:bold;font-size:1.1em;color:#1565C0;">D0-{r+1}</span>
                    <span style="background:#1976D2;color:white;padding:2px 8px;border-radius:4px;font-size:0.8em;">{shift_name}</span>
                    <span style="color:#666;font-size:0.85em;">지그교체: <b>{jig_changes[r]}</b></span>
                </div>
                {hanger_html}
                <div style="display:flex;flex-wrap:wrap;gap:4px;margin-top:5px;">{detail_html}</div>
            </div>'''

    # D0→D+1 전환
    html += f'''
            <div style="margin:15px 0;padding:10px;background:#FFE082;border-radius:8px;text-align:center;">
                <strong>▶ D0→D+1 전환</strong> | 지그교체: <b>{d1.get('start_jig_change', 0)}</b>
            </div>'''

    # D+1
    html += '<h3 style="margin-top:20px;color:#2E7D32;">D+1 생산계획</h3>'
    prev_positions_d1 = prev_positions_d0
    prev_grp_colors_d1 = prev_grp_colors_d0
    for r in range(10):
        shift_name = '주간' if r < 5 else '야간'
        shift_bg = '#E8F5E9' if r < 5 else '#F1F8E9'
        detail_html = format_rotation_items_html(items, r, 'prod1', templates_d1, jig_orders_d1)
        hanger_html, curr_positions, curr_grp_colors = format_hanger_positions_html(
            templates_d1, jig_orders_d1, r, prev_positions_d1,
            items=items, prod_key='prod1', prev_grp_colors=prev_grp_colors_d1)
        prev_positions_d1 = curr_positions
        prev_grp_colors_d1 = curr_grp_colors

        html += f'''
            <div style="margin:8px 0;padding:10px;background:{shift_bg};border-radius:8px;">
                <div style="display:flex;align-items:center;gap:15px;margin-bottom:8px;">
                    <span style="font-weight:bold;font-size:1.1em;color:#2E7D32;">D+1-{r+1}</span>
                    <span style="background:#388E3C;color:white;padding:2px 8px;border-radius:4px;font-size:0.8em;">{shift_name}</span>
                    <span style="color:#666;font-size:0.85em;">지그교체: <b>{jig_changes_d1[r]}</b></span>
                </div>
                {hanger_html}
                <div style="display:flex;flex-wrap:wrap;gap:4px;margin-top:5px;">{detail_html}</div>
            </div>'''

    # D+1→D+2 전환
    html += f'''
            <div style="margin:15px 0;padding:10px;background:#FFE082;border-radius:8px;text-align:center;">
                <strong>▶ D+1→D+2 전환</strong> | 지그교체: <b>{d2.get('start_jig_change', 0)}</b>
            </div>'''

    # D+2
    html += '<h3 style="margin-top:20px;color:#E65100;">D+2 생산계획</h3>'
    prev_positions_d2 = prev_positions_d1
    prev_grp_colors_d2 = prev_grp_colors_d1
    for r in range(10):
        shift_name = '주간' if r < 5 else '야간'
        shift_bg = '#FFF3E0' if r < 5 else '#FBE9E7'
        detail_html = format_rotation_items_html(items, r, 'prod2', templates_d2_box, jig_orders_d2)
        hanger_html, curr_positions, curr_grp_colors = format_hanger_positions_html(
            templates_d2_box, jig_orders_d2, r, prev_positions_d2,
            items=items, prod_key='prod2', prev_grp_colors=prev_grp_colors_d2)
        prev_positions_d2 = curr_positions
        prev_grp_colors_d2 = curr_grp_colors

        html += f'''
            <div style="margin:8px 0;padding:10px;background:{shift_bg};border-radius:8px;">
                <div style="display:flex;align-items:center;gap:15px;margin-bottom:8px;">
                    <span style="font-weight:bold;font-size:1.1em;color:#E65100;">D+2-{r+1}</span>
                    <span style="background:#F57C00;color:white;padding:2px 8px;border-radius:4px;font-size:0.8em;">{shift_name}</span>
                    <span style="color:#666;font-size:0.85em;">지그교체: <b>{jig_changes_d2_box[r]}</b></span>
                </div>
                {hanger_html}
                <div style="display:flex;flex-wrap:wrap;gap:4px;margin-top:5px;">{detail_html}</div>
            </div>'''

    # =============================================
    # 3일치 아이템별 수요/생산/재고 상세 테이블
    # =============================================
    html += '''
        <h2 style="margin-top:40px;">3일치 아이템별 수요/생산/재고 상세</h2>
        <div class="card" style="overflow-x:auto;">
        <table class="jig-table" style="font-size:0.75em;">
        <thead>
            <tr>
                <th rowspan="2">차종</th>
                <th rowspan="2">아이템</th>
                <th rowspan="2">세부</th>
                <th rowspan="2">컬러</th>
                <th rowspan="2">지그</th>
                <th rowspan="2">기초재고</th>'''

    # D0 헤더
    for i in range(10):
        html += f'<th colspan="3" style="background:#1976D2;">D0-{i+1}</th>'
    html += '<th rowspan="2" style="background:#0D47A1;">D0말</th>'

    # D+1 헤더
    for i in range(10):
        html += f'<th colspan="3" style="background:#388E3C;">D+1-{i+1}</th>'
    html += '<th rowspan="2" style="background:#1B5E20;">D+1말</th>'

    # D+2 헤더
    for i in range(10):
        html += f'<th colspan="3" style="background:#F57C00;">D+2-{i+1}</th>'
    html += '<th rowspan="2" style="background:#E65100;">D+2말</th>'

    html += '</tr><tr>'
    for day in ['D0', 'D+1', 'D+2']:
        bg = '#1976D2' if day == 'D0' else '#388E3C' if day == 'D+1' else '#F57C00'
        for i in range(10):
            html += f'<th style="background:{bg};font-size:0.8em;">수</th>'
            html += f'<th style="background:{bg};font-size:0.8em;">생</th>'
            html += f'<th style="background:{bg};font-size:0.8em;">재</th>'
    html += '</tr></thead><tbody>'

    # 아이템별 행
    for x in items:
        ct = x['ct'].replace('\n', ' ')[:10]
        it = x['it'].replace('\n', ' ')[:8]
        det = (x['det'] or '-').replace('\n', ' ')[:6]
        clr = x['clr'][:5] if x['clr'] else '-'
        grp = x['grp'] or '-'

        html += f'<tr><td>{ct}</td><td>{it}</td><td>{det}</td><td>{clr}</td><td>{grp}</td><td>{x["stk"]}</td>'

        # D0
        stk = x['stk']
        for i in range(10):
            dem, prd = x['d0'][i], x['prod'][i]
            stk = stk - dem + prd
            stk_style = 'color:#D32F2F;font-weight:bold;background:#FFEBEE;' if stk < 0 else ''
            html += f'<td>{dem}</td><td style="color:#1976D2;">{prd}</td><td style="{stk_style}">{stk}</td>'
        html += f'<td style="font-weight:bold;">{x["cur"]}</td>'

        # D+1
        stk = x['cur']
        for i in range(10):
            dem, prd = x['d1'][i], x['prod1'][i]
            stk = stk - dem + prd
            stk_style = 'color:#D32F2F;font-weight:bold;background:#FFEBEE;' if stk < 0 else ''
            html += f'<td>{dem}</td><td style="color:#388E3C;">{prd}</td><td style="{stk_style}">{stk}</td>'
        html += f'<td style="font-weight:bold;">{x["cur1"]}</td>'

        # D+2
        stk = x['cur1']
        for i in range(10):
            dem, prd = x['d2'][i], x['prod2'][i]
            stk = stk - dem + prd
            stk_style = 'color:#D32F2F;font-weight:bold;background:#FFEBEE;' if stk < 0 else ''
            html += f'<td>{dem}</td><td style="color:#F57C00;">{prd}</td><td style="{stk_style}">{stk}</td>'
        html += f'<td style="font-weight:bold;">{x["cur2"]}</td>'

        html += '</tr>'

    html += '</tbody></table></div>'

    # =============================================
    # D0 생산만으로 3일 재고 전망 테이블
    # =============================================
    html += '''
        <h2 style="margin-top:40px;">D0 생산만으로 3일 재고 전망 (부족 시 빨간색)</h2>
        <p style="color:#666;font-size:0.9em;">D0 생산량만 반영했을 때 D0/D+1/D+2 수요를 감당할 수 있는지 시뮬레이션</p>
        <div class="card" style="overflow-x:auto;">
        <table class="jig-table" style="font-size:0.75em;">
        <thead>
            <tr>
                <th rowspan="2">차종</th>
                <th rowspan="2">아이템</th>
                <th rowspan="2">세부</th>
                <th rowspan="2">컬러</th>
                <th rowspan="2">지그</th>
                <th rowspan="2">기초</th>
                <th rowspan="2">D0생산</th>'''

    # D0 수요/생산/재고 (3열)
    for i in range(10):
        html += f'<th colspan="3" style="background:#1976D2;">D0-{i+1}</th>'
    # D+1 수요/재고 (2열)
    for i in range(10):
        html += f'<th colspan="2" style="background:#388E3C;">D+1-{i+1}</th>'
    # D+2 수요/재고 (2열)
    for i in range(10):
        html += f'<th colspan="2" style="background:#F57C00;">D+2-{i+1}</th>'
    html += '<th rowspan="2">최종</th>'

    html += '</tr><tr>'
    # D0: 수/생/재
    for i in range(10):
        html += f'<th style="background:#1976D2;font-size:0.8em;">수</th>'
        html += f'<th style="background:#1976D2;font-size:0.8em;">생</th>'
        html += f'<th style="background:#1976D2;font-size:0.8em;">재</th>'
    # D+1: 수/재
    for i in range(10):
        html += f'<th style="background:#388E3C;font-size:0.8em;">수</th>'
        html += f'<th style="background:#388E3C;font-size:0.8em;">재</th>'
    # D+2: 수/재
    for i in range(10):
        html += f'<th style="background:#F57C00;font-size:0.8em;">수</th>'
        html += f'<th style="background:#F57C00;font-size:0.8em;">재</th>'
    html += '</tr></thead><tbody>'

    # 부족 아이템 먼저 정렬
    def get_min_stock_d0_only(x):
        stk = x['stk']
        d0_prod = sum(x['prod'])
        min_stk = stk + d0_prod
        for i in range(10):
            stk = stk - x['d0'][i] + x['prod'][i]
            min_stk = min(min_stk, stk)
        for i in range(10):
            stk = stk - x['d1'][i]
            min_stk = min(min_stk, stk)
        for i in range(10):
            stk = stk - x['d2'][i]
            min_stk = min(min_stk, stk)
        return min_stk

    items_sorted = sorted(items, key=get_min_stock_d0_only)

    for x in items_sorted:
        ct = x['ct'].replace('\n', ' ')[:10]
        it = x['it'].replace('\n', ' ')[:8]
        det = (x['det'] or '-').replace('\n', ' ')[:6]
        clr = x['clr'][:5] if x['clr'] else '-'
        grp = x['grp'] or '-'
        d0_prod_total = sum(x['prod'])

        has_shortage = get_min_stock_d0_only(x) < 0
        row_style = 'background:#FFF3E0;' if has_shortage else ''

        html += f'<tr style="{row_style}"><td>{ct}</td><td>{it}</td><td>{det}</td><td>{clr}</td><td>{grp}</td>'
        html += f'<td>{x["stk"]}</td><td style="color:#1976D2;font-weight:bold;">{d0_prod_total}</td>'

        # D0 (실제 생산 반영) - 수요/생산/재고
        stk = x['stk']
        for i in range(10):
            dem = x['d0'][i]
            prod = x['prod'][i]
            stk = stk - dem + prod
            stk_style = 'color:#D32F2F;font-weight:bold;background:#FFCDD2;' if stk < 0 else ''
            prod_style = 'color:#1976D2;font-weight:bold;' if prod > 0 else 'color:#ccc;'
            prod_text = prod if prod > 0 else '-'
            html += f'<td>{dem}</td><td style="{prod_style}">{prod_text}</td><td style="{stk_style}">{stk}</td>'

        # D+1 (D0 생산만, D+1 생산 없음)
        for i in range(10):
            dem = x['d1'][i]
            stk = stk - dem
            stk_style = 'color:#D32F2F;font-weight:bold;background:#FFCDD2;' if stk < 0 else ''
            html += f'<td>{dem}</td><td style="{stk_style}">{stk}</td>'

        # D+2 (D0 생산만)
        for i in range(10):
            dem = x['d2'][i]
            stk = stk - dem
            stk_style = 'color:#D32F2F;font-weight:bold;background:#FFCDD2;' if stk < 0 else ''
            html += f'<td>{dem}</td><td style="{stk_style}">{stk}</td>'

        final_style = 'color:#D32F2F;font-weight:bold;background:#FFCDD2;' if stk < 0 else 'font-weight:bold;'
        html += f'<td style="{final_style}">{stk}</td></tr>'

    html += '</tbody></table></div>'

    # =============================================
    # 오늘 생산 없을 때 재고 부족량 테이블
    # =============================================
    html += '''
        <h2 style="margin-top:40px;">오늘 생산 없을 때 재고 부족량</h2>
        <p style="color:#666;font-size:0.9em;">D0 생산이 전혀 없다고 가정할 때 각 시점의 재고 부족량 (부족 아이템만 표시)</p>
        <div class="card" style="overflow-x:auto;">
        <table class="jig-table" style="font-size:0.8em;">
        <thead>
            <tr>
                <th>차종</th>
                <th>아이템</th>
                <th>세부</th>
                <th>컬러</th>
                <th>지그</th>
                <th>기초재고</th>
                <th>D0 총수요</th>
                <th>D0말 재고<br>(생산無)</th>
                <th>D+1 오전<br>최소재고</th>
                <th>D+1말 재고</th>
                <th>최대부족량</th>
                <th>부족시점</th>
            </tr>
        </thead>
        <tbody>'''

    # 생산 없을 때 부족 계산
    no_prod_shortages = []
    for x in items:
        stk = x['stk']
        d0_total = sum(x['d0'])
        d1_morning = sum(x['d1'][:5])  # D+1 오전
        d1_total = sum(x['d1'])

        # D0말 재고 (생산 없음)
        d0_end_no_prod = stk - d0_total

        # D+1 오전 최소 재고
        d1_morning_min = d0_end_no_prod
        temp_stk = d0_end_no_prod
        for i in range(5):
            temp_stk -= x['d1'][i]
            d1_morning_min = min(d1_morning_min, temp_stk)

        # D+1말 재고
        d1_end_no_prod = d0_end_no_prod - d1_total

        # 최대 부족량 및 시점 찾기
        max_shortage = 0
        shortage_time = ''
        temp_stk = stk
        for i in range(10):
            temp_stk -= x['d0'][i]
            if temp_stk < max_shortage:
                max_shortage = temp_stk
                shortage_time = f'D0-{i+1}'
        for i in range(10):
            temp_stk -= x['d1'][i]
            if temp_stk < max_shortage:
                max_shortage = temp_stk
                shortage_time = f'D+1-{i+1}'

        if max_shortage < 0:
            no_prod_shortages.append({
                'x': x,
                'd0_total': d0_total,
                'd0_end': d0_end_no_prod,
                'd1_morning_min': d1_morning_min,
                'd1_end': d1_end_no_prod,
                'max_shortage': -max_shortage,
                'shortage_time': shortage_time
            })

    # 부족량 큰 순 정렬
    no_prod_shortages.sort(key=lambda s: -s['max_shortage'])

    for s in no_prod_shortages:
        x = s['x']
        ct = x['ct'].replace('\n', ' ')[:12]
        it = x['it'].replace('\n', ' ')[:10]
        det = (x['det'] or '-').replace('\n', ' ')[:8]
        clr = x['clr'][:6] if x['clr'] else '-'
        grp = x['grp'] or '-'

        d0_end_style = 'color:#D32F2F;font-weight:bold;background:#FFCDD2;' if s['d0_end'] < 0 else ''
        d1m_style = 'color:#D32F2F;font-weight:bold;background:#FFCDD2;' if s['d1_morning_min'] < 0 else ''
        d1_end_style = 'color:#D32F2F;font-weight:bold;background:#FFCDD2;' if s['d1_end'] < 0 else ''

        html += f'''<tr>
            <td>{ct}</td><td>{it}</td><td>{det}</td><td>{clr}</td><td>{grp}</td>
            <td>{x['stk']}</td>
            <td>{s['d0_total']}</td>
            <td style="{d0_end_style}">{s['d0_end']}</td>
            <td style="{d1m_style}">{s['d1_morning_min']}</td>
            <td style="{d1_end_style}">{s['d1_end']}</td>
            <td style="color:#D32F2F;font-weight:bold;">{s['max_shortage']}</td>
            <td style="color:#D32F2F;">{s['shortage_time']}</td>
        </tr>'''

    if not no_prod_shortages:
        html += '<tr><td colspan="12" style="text-align:center;color:#388E3C;">부족 아이템 없음</td></tr>'

    html += '</tbody></table></div>'

    html += '''
        </div>
    </div>
</body>
</html>'''

    return html


def save_csv(items, filename='production_plan_v8.csv'):
    """CSV 저장"""
    import csv
    with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        header = ['차종', '아이템', '세부', '컬러', '지그', '기초재고']
        for i in range(10):
            header.extend([f'D0_{i+1}수요', f'D0_{i+1}생산'])
        header.append('D0기말')
        for i in range(10):
            header.extend([f'D1_{i+1}수요', f'D1_{i+1}생산'])
        header.append('D1기말')
        for i in range(10):
            header.extend([f'D2_{i+1}수요', f'D2_{i+1}생산'])
        header.append('D2기말')
        writer.writerow(header)

        for x in items:
            row = [x['ct'], x['it'], x['det'], x['clr'], x['grp'], x['stk']]
            for i in range(10):
                row.extend([x['d0'][i], x['prod'][i]])
            row.append(x['cur'])
            for i in range(10):
                row.extend([x['d1'][i], x['prod1'][i]])
            row.append(x['cur1'])
            for i in range(10):
                row.extend([x['d2'][i], x['prod2'][i]])
            row.append(x['cur2'])
            writer.writerow(row)


if __name__ == '__main__':
    print("데이터 로드...")
    items = load_data()
    print(f"{len(items)}개 아이템")

    print("스케줄링...")
    result = schedule(items)

    print("HTML 리포트 생성...")
    html = generate_html_report(items, result)
    with open('production_report.html', 'w', encoding='utf-8') as f:
        f.write(html)

    print("CSV 저장...")
    save_csv(items)

    print("\n완료!")
    print("  => production_report.html")
    print("  => production_plan_v8.csv")
